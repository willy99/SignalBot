from dics.deserter_xls_dic import VOC_NAMES, COLUMN_DESERTION_PLACE, MIL_UNITS
from domain.person_filter import PersonSearchFilter
from gui.services.auth_manager import AuthManager
from gui.views.person.search_view import results_ui  # Використовуємо існуючу таблицю
from nicegui import ui
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
import io

async def enhanced_search_page(person_ctrl, auth_manager: AuthManager):
    current_results = []
    ui_options = person_ctrl.get_column_options()

    # Формуємо опції для селекта ВОС
    voc_options = {'Всі': 'Всі категорії'}
    voc_options.update(VOC_NAMES)

    def on_voc_change(e):
        if len(e.value) > 1 and 'Всі' in e.value:
            if e.value[0] == 'Всі':
                voc_select.value = [v for v in e.value if v != 'Всі']
            else:
                voc_select.value = ['Всі']

    def export_to_excel():
        if not current_results:
            ui.notify('Немає даних для експорту', type='warning')
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Результати аналітики"

        # Стилі
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="EEEEEE")
        center_align = Alignment(horizontal="center", vertical="center")

        # Визначаємо колонки
        cols_config = [
            ('ВЧ', 'mil_unit'),
            ('ПІБ', 'name'),
            ('Звання', 'title2'),
            ('РНОКПП', 'rnokpp'),
            ('Підрозділ', 'subunit'),
            ('Дата СЗЧ', 'desertion_date'),
            ('Звідки', 'desertion_place'),
            ('Знайдена посада (ВОС)', 'matched_voc_info'),
            ('Статус', 'review_status')
        ]

        # 1. Записуємо хедер
        for col_num, (header_text, _) in enumerate(cols_config, start=1):
            cell = ws.cell(row=1, column=col_num, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # 2. Записуємо дані
        for row_num, person in enumerate(current_results, start=2):
            for col_num, (_, attr_name) in enumerate(cols_config, start=1):
                # Отримуємо значення (використовуємо getattr для віртуальних полів)
                val = getattr(person, attr_name, "")

                # Обробка дат, щоб вони не перетворювалися на текст
                if hasattr(val, 'isoformat'):
                    val = val.strftime('%d.%m.%Y')

                cell = ws.cell(row=row_num, column=col_num, value=val)
                cell.border = thin_border

        # 3. Автопідбір ширини колонок
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        # 4. Збереження у буфер та вивантаження
        output = io.BytesIO()
        wb.save(output)
        ui.download(output.getvalue(), filename=f"report_szch_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        ui.notify('Файл сформовано')

    async def do_event_search():
        nonlocal current_results
        # Валідація: хоча б один фільтр має бути обраний
        if not any([date_from.value, date_to.value,
                    place_filter.value != 'Всі',
                    voc_select.value != 'Всі']):
            ui.notify('Оберіть хоча б один параметр для фільтрації', type='warning')
            return
        selected_vocs = voc_select.value
        actual_vocs = None if 'Всі' in selected_vocs or not selected_vocs else selected_vocs

        results_container.clear()
        with results_container:
            ui.spinner(size='lg').classes('mt-10')

        # Формуємо фільтр для контролера
        search_filter = PersonSearchFilter(
            des_date_from=date_from.value,
            des_date_to=date_to.value,
            desertion_place=None if place_filter.value == 'Всі' else place_filter.value,
            voc_codes=actual_vocs,
            mil_unit=sheet_select.value
        )

        try:
            # Викликаємо метод пошуку подій
            data = await auth_manager.execute(
                person_ctrl.search,
                auth_manager.get_current_context(),
                search_filter
            )

            current_results = data

            results_container.clear()

            if not data:
                ui.notify('Подій за такими критеріями не знайдено', type='negative')
                return

            export_btn.set_visibility(len(data) > 0)

            with results_container:
                # Використовуємо стандартний results_ui для відображення таблиці
                # refresh_callback дозволяє оновити пошук після редагування запису
                refresh_cb = lambda: ui.timer(0, do_event_search, once=True)
                results_ui(data, person_ctrl, auth_manager, refresh_callback=refresh_cb)

        except Exception as e:
            results_container.clear()
            ui.notify(f'Помилка аналітики: {e}', type='negative')

    # --- UI LAYOUT ---
    with ui.column().classes('w-full items-center p-4'):
        with ui.row().classes('items-center gap-3 mb-4'):
            ui.icon('analytics', size='lg', color='indigo')
            ui.label('Розширений пошук та аналітика подій').classes('text-h4')

        with ui.card().classes('w-full max-w-6xl p-6 shadow-lg border-t-4 border-indigo-500'):
            with ui.row().classes('w-full items-center gap-4'):
                sheet_select = ui.select(options=MIL_UNITS, value=MIL_UNITS[0], label='ВЧ').classes('w-32')

                date_from = ui.input('СЗЧ з (дата)').props('type=date clearable').classes('w-40')
                date_to = ui.input('СЗЧ по (дата)').props('type=date clearable').classes('w-40')

                places = ['Всі'] + [p for p in ui_options.get(COLUMN_DESERTION_PLACE, []) if p]
                place_filter = ui.select(options=places, value='Всі', label='Звідки (локація)').classes('flex-grow')

            with ui.row().classes('w-full items-center gap-4 mt-4'):
                voc_select = ui.select(
                    options=voc_options,
                    value=['Всі'],  # Початкове значення списком
                    label='Спеціалізація',
                    multiple=True  # Дозволяємо множинний вибір
                ).classes('flex-grow').props('use-chips stack-label')
                voc_select.on_value_change(on_voc_change)

                ui.button('Сформувати звіт', icon='bolt', on_click=do_event_search) \
                    .classes('h-14 bg-indigo-600 text-white px-8') \
                    .props('elevated')
                export_btn = ui.button('Завантажити Excel', icon='download', on_click=export_to_excel) \
                    .classes('h-14 bg-green-700 text-white px-8')
                export_btn.set_visibility(False)

        # Контейнер для результатів
        results_container = ui.column().classes('w-full items-center mt-6')

    # Додаємо гарячу клавішу Enter для запуску пошуку
    ui.on('keydown.enter', do_event_search)