from nicegui import ui, run

import utils.utils
from gui.controllers.person_controller import PersonController
from gui.controllers.report_controller import ReportController
from gui.services.auth_manager import AuthManager
from domain.person_filter import PersonSearchFilter
from dics.deserter_xls_dic import REVIEW_STATUS_NOT_ASSIGNED, REVIEW_STATUS_ASSIGNED, REVIEW_STATUS_CLOSED, COLUMN_INSERT_DATE
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from datetime import datetime

def render_place_report_page(report_ctrl:ReportController, person_ctrl: PersonController, auth_manager: AuthManager):
    state = {'rows': [], 'columns': []}
    year_options = person_ctrl.get_column_options().get(COLUMN_INSERT_DATE, [])
    year_options = sorted([str(y) for y in year_options], reverse=True)
    with ui.column().classes('w-full items-center p-4'):
        ui.label('Аналітика: Загальний стан справ').classes('text-h4 mb-6')

        with ui.row().classes('w-full max-w-4xl justify-center items-center gap-4 mb-6'):
            # Додамо фільтр по роках, якщо треба (опціонально)
            year_select = ui.select(year_options, label='Рік').classes('w-32')

            generate_btn = ui.button('Сформувати', icon='analytics',
                                     on_click=lambda: do_report()) \
                .props('elevated color="primary"')

            export_btn = ui.button('Excel', icon='download', color='green',
                                   on_click=lambda: export_place_report_to_excel(state['rows'], state['columns'])) \
                .props('elevated').bind_visibility_from(state, 'rows', backward=lambda r: len(r) > 0)

        results_container = ui.column().classes('w-full items-center mt-2')

    async def do_report():
        results_container.clear()
        with results_container:
            ui.spinner(size='xl')

        try:
            # Створюємо фільтр (можна розширити)
            filt = PersonSearchFilter()
            if year_select.value:
                filt.des_year = [year_select.value]

            # Викликаємо твій новий метод з контролера
            data = await auth_manager.execute(report_ctrl.get_general_state_report, auth_manager.get_current_context(), filt)

            results_container.clear()
            if not data:
                ui.notify('Дані за вказаними критеріями відсутні', type='warning')
                return

            with results_container:
                rows, columns = build_place_table(data)
                state['rows'] = rows
                state['columns'] = columns

        except Exception as e:
            ui.notify(f'Помилка: {e}', type='negative')


def build_place_table(data):
    rows = []

    # Створюємо рядок "РАЗОМ"
    grand_total = {
        'place': 'РАЗОМ ПО ВЧ',
        REVIEW_STATUS_NOT_ASSIGNED: 0,
        REVIEW_STATUS_ASSIGNED: 0,
        REVIEW_STATUS_CLOSED: 0,
        'term_under_10': 0,
        'term_10_30': 0,
        'term_over_30': 0,
        'total': 0,
        'is_grand_total': True
    }

    # Перетворюємо dict у список рядків для таблиці
    for place, stats in data.items():
        row = {'place': place, 'is_grand_total': False, **stats}
        rows.append(row)

        for key in grand_total:
            if key not in ['place', 'is_grand_total']:
                grand_total[key] += stats.get(key, 0)

    rows.sort(key=lambda x: x['total'], reverse=True)
    rows.append(grand_total)

    columns = [
        {'name': 'place', 'label': 'Обставини (звідки)', 'field': 'place', 'align': 'left'},
        {'name': 'not_assigned', 'label': 'Не призначено', 'field': REVIEW_STATUS_NOT_ASSIGNED},
        {'name': 'assigned', 'label': 'Призначено', 'field': REVIEW_STATUS_ASSIGNED},
        {'name': 'closed', 'label': 'Закрито', 'field': REVIEW_STATUS_CLOSED},
        {'name': 'u10', 'label': 'до 10 діб', 'field': 'term_under_10', 'headerClasses': 'bg-orange-50'},
        {'name': 'u30', 'label': '10-30 діб', 'field': 'term_10_30', 'headerClasses': 'bg-orange-50'},
        {'name': 'o30', 'label': '> 30 діб', 'field': 'term_over_30', 'headerClasses': 'bg-red-50'},
        {'name': 'total', 'label': 'Всього СЗЧ', 'field': 'total', 'headerClasses': 'bg-blue-100 font-bold'},
    ]

    table = ui.table(columns=columns, rows=rows, row_key='place').classes('w-full max-w-5xl')
    table.props('bordered separator=cell flat dense')

    # Шапка з об'єднанням стовпців (як ти любиш)
    table.add_slot('header', '''
        <q-tr>
            <q-th colspan="1" class="bg-grey-2 text-bold">Місце</q-th>
            <q-th colspan="3" class="bg-indigo-1 text-bold">Статус розслідування</q-th>
            <q-th colspan="3" class="bg-orange-1 text-bold">Тривалість СЗЧ (календарних діб)</q-th>
            <q-th colspan="1" class="bg-blue-2 text-bold text-subtitle2">Загалом</q-th>
        </q-tr>
        <q-tr>
            <q-th v-for="col in props.cols" :key="col.name" :props="props">
                {{ col.label }}
            </q-th>
        </q-tr>
    ''')

    # Кастомні рядки з підсвіткою "РАЗОМ"
    table.add_slot('body', '''
        <q-tr :props="props" :class="props.row.is_grand_total ? 'bg-orange-1 font-bold' : ''">
            <q-td v-for="col in props.cols" :key="col.name" :props="props">
                <template v-if="col.name === 'place' && props.row.is_grand_total">
                    <q-icon name="summarize" color="primary" /> {{ col.value }}
                </template>
                <template v-else>
                    {{ col.value }}
                </template>
            </q-td>
        </q-tr>
    ''')

    return rows, columns


def export_place_report_to_excel(rows, columns):
    if not rows:
        ui.notify('Немає даних для експорту', type='warning')
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Звіт за обставинами"

    # --- СТИЛІ ---
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="EEEEEE")  # Сірий для шапки
    total_fill = PatternFill("solid", fgColor="FFE0B2")  # Помаранчевий для "РАЗОМ"
    blue_fill = PatternFill("solid", fgColor="E3F2FD")  # Блакитний для колонки "Всього"

    # --- 1. ПЕРШИЙ РЯДОК ШАПКИ (Групування) ---
    # Об'єднуємо комірки згідно з логікою build_place_table
    # A1: Місце, B1-D1: Статус, E1-G1: Тривалість, H1: Загалом

    def setup_header_cell(cell_range, value):
        ws.merge_cells(cell_range)
        cell = ws[cell_range.split(':')[0]]
        cell.value = value
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = header_fill
        # Проставляємо межі для всіх об'єднаних комірок
        for row in ws[cell_range]:
            for c in row:
                c.border = thin_border

    setup_header_cell('A1:A1', 'Обставини (звідки)')
    setup_header_cell('B1:D1', 'Статус розслідування')
    setup_header_cell('E1:G1', 'Тривалість СЗЧ (діб)')
    setup_header_cell('H1:H1', 'Всього СЗЧ')

    # --- 2. ДРУГИЙ РЯДОК ШАПКИ (Назви колонок) ---
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=idx, value=col['label'])
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align
        cell.fill = header_fill

    # --- 3. ЗАПОВНЕННЯ ДАНИМИ ---
    # Починаємо з 3-го рядка
    for r_idx, row_data in enumerate(rows, start=3):
        is_grand_total = row_data.get('is_grand_total', False)

        for c_idx, col_def in enumerate(columns, start=1):
            val = row_data.get(col_def['field'])
            cell = ws.cell(row=r_idx, column=c_idx, value=val)

            # Базові стилі для кожної клітинки
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center") if c_idx > 1 else Alignment(horizontal="left")

            # Підсвітка останнього рядка "РАЗОМ"
            if is_grand_total:
                cell.fill = total_fill
                cell.font = bold_font

            # Підсвітка останньої колонки "Всього"
            elif col_def['name'] == 'total':
                cell.fill = blue_fill
                cell.font = bold_font

    # --- 4. НАЛАШТУВАННЯ ВИГЛЯДУ ---
    ws.column_dimensions['A'].width = 40  # Ширша колонка для назв місць
    for col_let in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_let].width = 15

    # Фіксація шапки (перші 2 рядки)
    ws.freeze_panes = 'A3'

    # --- 5. ЗАВАНТАЖЕННЯ ---
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    ui.download(buffer.getvalue(), filename='Загальний Стан (' + utils.utils.format_to_excel_date(datetime.now()) + ').xlsx')
    ui.notify('Excel-файл сформовано', type='positive')
