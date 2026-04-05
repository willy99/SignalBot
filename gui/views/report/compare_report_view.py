"""
compare_report_view.py
======================
Порівняння зовнішнього Excel-файлу з основною базою СЗЧ.

Step 1 — Маппінг полів:
  - Завантаження xlsx-файлу
  - Відображення колонок файлу (ліворуч) і GENERAL_FIELDS бази (праворуч)
  - Drag-and-drop маппінг + інтелектуальний авто-маппінг

Step 2 — Вибір полів для порівняння + перенесення даних:
  - Чекбокси для полів файлу і бази
  - Над кожним стовпцем файлу — комбобокс «перенести в поле бази»
  - Кнопка «Виправити» там де поле бази порожнє, а у файлі є значення
  - «Зберегти дані» → PersonController.save_persons (partial_update=True)
  - Кнопка «Завантажити xlsx» → export результатів
"""

import io
from datetime import datetime, date

import openpyxl
from nicegui import ui, events

import config
from dics.deserter_xls_dic import (
    GENERAL_FIELDS,
    COLUMN_NAME,
    COLUMN_ID_NUMBER,
    COLUMN_BIRTHDAY,
    COLUMN_INCREMENTAL,
    COLUMN_MIL_UNIT,
    NA,
)
from domain.person import Person
from gui.services.auth_manager import AuthManager

# ---------------------------------------------------------------------------
# Ключові поля (виділяємо ★ в UI)
# ---------------------------------------------------------------------------
KEY_FIELDS = {COLUMN_NAME, COLUMN_ID_NUMBER, COLUMN_BIRTHDAY}

# ---------------------------------------------------------------------------
# Авто-маппінг
# ---------------------------------------------------------------------------
_AUTO_MATCH: dict[str, list[str]] = {
    COLUMN_NAME:      ['піб', 'pib', 'прізвище', 'фіо', 'name', 'fullname', "ім'я"],
    COLUMN_ID_NUMBER: ['рнокпп', 'rnokpp', 'інн', 'inn', 'ікн', 'taxid', 'код платника'],
    COLUMN_BIRTHDAY:  ['народження', 'birthday', 'birth', 'дн', 'дата нар'],
    'Дата СЗЧ':               ['сзч', 'szch', 'дата сзч', 'дата залишення'],
    'Військове звання':       ['звання', 'rank', 'чин'],
    'Підрозділ':              ['підрозділ', 'підрозд', 'unit', 'субюніт'],
    'Адреса проживання':      ['адрес', 'address', 'місце проживання'],
    '№ телефону':             ['телефон', 'phone', 'mobile', 'моб'],
}


def _smart_map(upload_cols: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for gen_field, keywords in _AUTO_MATCH.items():
        if gen_field not in GENERAL_FIELDS:
            continue
        for col in upload_cols:
            if any(kw in col.lower().strip() for kw in keywords):
                mapping[gen_field] = col
                break
    return mapping


# ---------------------------------------------------------------------------
# Утиліти дат
# ---------------------------------------------------------------------------

def _parse_date_for_sort(date_str: str) -> str:
    """
    Перетворює «dd.mm.YYYY» → «YYYY-MM-DD» для лексичного сортування.
    Повертає '' якщо розпарсити не вдається.
    """
    if not date_str or len(date_str) < 8:
        return ''
    try:
        return datetime.strptime(date_str, config.EXCEL_DATE_FORMAT).strftime('%Y-%m-%d')
    except ValueError:
        return ''


# ---------------------------------------------------------------------------
# Головна функція рендеру
# ---------------------------------------------------------------------------

def render_compare_report_page(report_ctrl, person_ctrl, auth_manager: AuthManager):
    state: dict = {
        'file_bytes':   None,
        'upload_cols':  [],
        'mapping':      {},       # {gen_field: upload_col}
        'sel_upload':   [],       # порядок важливий
        'sel_general':  [],       # порядок важливий
        'pending_saves': {},      # {row_idx: Person} — черга на збереження
    }

    with ui.column().classes('w-full p-4 gap-2'):
        ui.label('Порівняння файлів').classes('text-h5 font-bold')
        ui.label(
            'Завантажте Excel-файл, налаштуйте маппінг полів і '
            'оберіть колонки для порівняння з основною базою СЗЧ.'
        ).classes('text-grey text-sm max-w-3xl')

    content = ui.column().classes('w-full px-4 pb-4 gap-4')

    def render_step1():
        content.clear()
        with content:
            _render_stepper(1)
            with ui.card().classes('w-full p-4'):
                ui.label('1. Завантажте файл для порівняння').classes(
                    'text-subtitle1 font-bold mb-2'
                )
                mapping_area = ui.column().classes('w-full')

                async def handle_upload(e: events.UploadEventArguments):
                    state['file_bytes'] = await e.file.read()
                    try:
                        wb = openpyxl.load_workbook(
                            io.BytesIO(state['file_bytes']), read_only=True, data_only=True
                        )
                        ws = wb.active
                        headers = [
                            str(cell).strip()
                            for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                            if cell is not None
                        ]
                        wb.close()
                        state['upload_cols'] = headers
                        state['mapping']     = _smart_map(headers)
                        ui.notify(
                            f'Завантажено: {e.file.name} — {len(headers)} колонок, '
                            f'авто-маппінг: {len(state["mapping"])} полів',
                            type='positive',
                        )
                        _render_mapping_ui(mapping_area, state, on_next=render_step2)
                    except Exception as ex:
                        ui.notify(f'Помилка читання файлу: {ex}', type='negative')

                ui.upload(
                    label='Оберіть .xlsx файл',
                    auto_upload=True, max_files=1,
                    on_upload=handle_upload,
                ).props('accept=.xlsx').classes('w-full max-w-lg')

    def render_step2():
        if not state['mapping']:
            ui.notify('Налаштуйте маппінг хоча б одного ключового поля', type='warning')
            return
        if not state['sel_upload']:
            for gf, uc in state['mapping'].items():
                if gf in {COLUMN_NAME, COLUMN_ID_NUMBER} and uc not in state['sel_upload']:
                    state['sel_upload'].append(uc)
        if not state['sel_general']:
            for gf in (COLUMN_NAME, COLUMN_ID_NUMBER):
                if gf in state['mapping'] and gf not in state['sel_general']:
                    state['sel_general'].append(gf)
        content.clear()
        with content:
            _render_stepper(2)
            _render_field_selection(
                state, report_ctrl, person_ctrl, auth_manager, on_back=render_step1
            )

    render_step1()


# ---------------------------------------------------------------------------
# Stepper
# ---------------------------------------------------------------------------

def _render_stepper(current: int):
    steps = ['Маппінг полів', 'Вибір та порівняння']
    with ui.row().classes('w-full items-center gap-0 mb-2'):
        for i, label in enumerate(steps, 1):
            done   = i < current
            active = i == current
            circle_cls = (
                'bg-primary text-white' if active else
                ('bg-green-500 text-white' if done else 'bg-grey-3 text-grey-7')
            )
            with ui.row().classes('items-center gap-2'):
                with ui.element('div').classes(
                    f'rounded-full w-7 h-7 flex items-center justify-center '
                    f'text-sm font-bold {circle_cls}'
                ):
                    ui.label(str(i))
                ui.label(label).classes(
                    'text-sm' + (' font-bold text-primary' if active else
                                 ' text-green-600' if done else ' text-grey')
                )
            if i < len(steps):
                ui.element('div').classes('flex-1 h-px bg-grey-3 mx-2').style('min-width:32px')


# ---------------------------------------------------------------------------
# Step 1: маппінг (drag-and-drop)
# ---------------------------------------------------------------------------

def _render_mapping_ui(container: ui.column, state: dict, on_next):
    container.clear()
    upload_cols = state['upload_cols']
    mapping     = state['mapping']
    badge_containers: dict[str, ui.element] = {}

    with container:
        with ui.card().classes('w-full p-4 mt-3'):
            ui.label('2. Налаштуйте маппінг полів').classes('text-subtitle1 font-bold')
            ui.label(
                'Перетягніть колонки зліва на відповідні поля бази справа. '
                'Поля зі ★ є ключовими для пошуку.'
            ).classes('text-grey text-sm mb-3')

            with ui.row().classes('w-full gap-4 items-start'):
                with ui.column().classes('gap-2').style('min-width:220px; flex:1'):
                    ui.label('Колонки вашого файлу').classes(
                        'text-caption text-grey font-bold uppercase'
                    )
                    for col in upload_cols:
                        (
                            ui.chip(col, icon='drag_indicator')
                            .classes(
                                'upload-col-chip cursor-grab bg-blue-1 text-blue-9 '
                                'w-full justify-start'
                            )
                            .props(f'data-col="{col}"')
                        )

                with ui.column().classes('gap-1').style('flex:1'):
                    ui.label('Поля основної бази СЗЧ').classes(
                        'text-caption text-grey font-bold uppercase'
                    )
                    for gen_field in GENERAL_FIELDS:
                        is_key = gen_field in KEY_FIELDS
                        with ui.row().classes(
                            'general-field-row w-full items-center gap-2 px-2 py-1 '
                            'rounded border border-grey-3 hover:border-primary transition-all'
                        ).props(f'data-field="{gen_field}"'):
                            if is_key:
                                ui.icon('star', color='orange', size='xs').classes('flex-shrink-0')
                            else:
                                ui.element('div').classes('w-4 flex-shrink-0')
                            ui.label(gen_field).classes('text-sm flex-1')
                            badge_cont = ui.element('div').classes('flex-shrink-0')
                            with badge_cont:
                                if mapping.get(gen_field):
                                    ui.badge(mapping[gen_field], color='primary').classes('text-xs')
                                else:
                                    ui.label('—').classes('text-grey text-xs')
                            badge_containers[gen_field] = badge_cont

            summary = ui.column().classes('w-full mt-3 p-3 bg-grey-1 rounded gap-1')
            _refresh_summary(summary, mapping)

            with ui.row().classes('w-full justify-end gap-3 mt-4'):
                ui.button(
                    'Скинути', icon='clear',
                    on_click=lambda: _clear_mapping(state, badge_containers, summary),
                ).props('flat')
                ui.button('Далі →', icon='arrow_forward', on_click=on_next).props(
                    'elevated color=primary'
                )

    def handle_mapped(e):
        args = e.args if hasattr(e, 'args') else {}
        gen_field  = args.get('gen_field', '')
        upload_col = args.get('upload_col', '')
        if not gen_field or not upload_col:
            return
        state['mapping'][gen_field] = upload_col
        badge_cont = badge_containers.get(gen_field)
        if badge_cont:
            badge_cont.clear()
            with badge_cont:
                ui.badge(upload_col, color='primary').classes('text-xs')
        _refresh_summary(summary, state['mapping'])

    ui.on('field_mapped', handle_mapped)
    ui.timer(0.25, _inject_dnd_js, once=True)


def _inject_dnd_js():
    ui.run_javascript('''
        (function setup_dnd() {
            const chips = document.querySelectorAll('.upload-col-chip');
            const rows  = document.querySelectorAll('.general-field-row');
            if (!chips.length || !rows.length) { setTimeout(setup_dnd, 150); return; }
            chips.forEach(el => {
                el.setAttribute('draggable', 'true');
                el.addEventListener('dragstart', function(ev) {
                    window._dndCol = this.getAttribute('data-col') || this.innerText.trim();
                    ev.dataTransfer.effectAllowed = 'copy';
                    this.style.opacity = '0.45';
                });
                el.addEventListener('dragend', function() { this.style.opacity = '1'; });
            });
            rows.forEach(el => {
                el.addEventListener('dragover', function(ev) {
                    ev.preventDefault();
                    ev.dataTransfer.dropEffect = 'copy';
                    this.style.background = '#e3f2fd';
                });
                el.addEventListener('dragleave', function() { this.style.background = ''; });
                el.addEventListener('drop', function(ev) {
                    ev.preventDefault();
                    this.style.background = '';
                    const genField  = this.getAttribute('data-field');
                    const uploadCol = window._dndCol;
                    if (genField && uploadCol)
                        emitEvent('field_mapped', {gen_field: genField, upload_col: uploadCol});
                });
            });
        })();
    ''')


def _refresh_summary(container, mapping):
    container.clear()
    if not mapping:
        with container:
            ui.label('Маппінг ще не налаштовано').classes('text-grey text-sm')
        return
    with container:
        ui.label('Поточний маппінг:').classes('text-caption text-grey')
        for gen_field, upload_col in mapping.items():
            star = '★ ' if gen_field in KEY_FIELDS else ''
            with ui.row().classes('gap-2 items-center'):
                ui.label(upload_col).classes('text-sm text-blue-9 font-medium')
                ui.icon('arrow_forward', size='xs').classes('text-grey')
                ui.label(f'{star}{gen_field}').classes('text-sm')


def _clear_mapping(state, badge_containers, summary):
    state['mapping'].clear()
    for gen_field, badge_cont in badge_containers.items():
        badge_cont.clear()
        with badge_cont:
            ui.label('—').classes('text-grey text-xs')
    _refresh_summary(summary, {})
    ui.notify('Маппінг скинуто', type='info')


# ---------------------------------------------------------------------------
# Step 2: вибір полів + порівняння
# ---------------------------------------------------------------------------

def _render_field_selection(state, report_ctrl, person_ctrl, auth_manager, on_back):
    upload_cols = state['upload_cols']
    sel_upload  = state['sel_upload']
    sel_general = state['sel_general']

    left_checks:  dict[str, ui.checkbox] = {}
    right_checks: dict[str, ui.checkbox] = {}

    with ui.card().classes('w-full p-4'):
        ui.label('3. Оберіть поля для порівняння').classes('text-subtitle1 font-bold')
        ui.label(
            'Позначте колонки файлу (ліворуч) і поля бази (праворуч). '
            'Порядок вибору = порядок колонок у звіті.'
        ).classes('text-grey text-sm mb-3')

        with ui.row().classes('w-full gap-6 items-start'):
            with ui.column().classes('gap-1').style('flex:1'):
                with ui.row().classes('w-full items-center justify-between mb-1'):
                    ui.label('Колонки файлу').classes('text-caption font-bold uppercase text-grey')
                    ui.button('Всі', icon='select_all', on_click=lambda: _select_all(
                        upload_cols, sel_upload, left_checks
                    )).props('flat dense size=xs')
                    ui.button('Жодного', icon='deselect', on_click=lambda: _deselect_all(
                        sel_upload, left_checks
                    )).props('flat dense size=xs')
                for col in upload_cols:
                    cb = ui.checkbox(
                        col, value=(col in sel_upload),
                        on_change=lambda e, c=col: _toggle(sel_upload, c, e.value, e.sender),
                    ).classes('text-sm')
                    left_checks[col] = cb

            with ui.column().classes('gap-1').style('flex:1'):
                with ui.row().classes('w-full items-center justify-between mb-1'):
                    ui.label('Поля бази СЗЧ').classes('text-caption font-bold uppercase text-grey')
                    ui.button('Всі', icon='select_all', on_click=lambda: _select_all(
                        list(GENERAL_FIELDS), sel_general, right_checks
                    )).props('flat dense size=xs')
                    ui.button('Жодного', icon='deselect', on_click=lambda: _deselect_all(
                        sel_general, right_checks
                    )).props('flat dense size=xs')
                for gen_field in GENERAL_FIELDS:
                    is_key = gen_field in KEY_FIELDS
                    label  = ('★ ' if is_key else '') + gen_field
                    cb = ui.checkbox(
                        label, value=(gen_field in sel_general),
                        on_change=lambda e, f=gen_field: _toggle(sel_general, f, e.value, e.sender),
                    ).classes('text-sm' + (' text-orange font-bold' if is_key else ''))
                    right_checks[gen_field] = cb

    compare_results = ui.column().classes('w-full mt-2')

    # Кнопки — Назад ліворуч, Завантажити + Зберегти + Порівняти праворуч
    with ui.row().classes('w-full justify-between items-center gap-3 mt-2'):
        ui.button('← Назад', icon='arrow_back', on_click=on_back).props('flat')
        with ui.row().classes('gap-2 items-center'):
            export_btn = ui.button(
                'Завантажити xlsx', icon='download',
                on_click=lambda: _export_compare(
                    state.get('_last_rows', []),
                    list(state['sel_upload']),
                    list(state['sel_general']),
                ),
            ).props('elevated color=green')
            export_btn.set_visibility(False)
            state['_export_btn'] = export_btn

            save_btn = ui.button(
                'Зберегти дані', icon='cloud_upload',
                on_click=lambda: ui.timer(
                    0,
                    lambda: _do_save(state, person_ctrl, auth_manager, save_btn),
                    once=True,
                ),
            ).props('elevated color=orange')
            save_btn.set_visibility(False)
            state['_save_btn'] = save_btn

            ui.button(
                'Порівняти', icon='compare_arrows',
                on_click=lambda: ui.timer(
                    0,
                    lambda: _do_compare(state, report_ctrl, auth_manager, compare_results),
                    once=True,
                ),
            ).props('elevated color=primary')


# ---------------------------------------------------------------------------
# Допоміжні функції вибору чекбоксів
# ---------------------------------------------------------------------------

def _toggle(sel_list, field, value, sender=None):
    if value:
        if sender is not None and not sender.value:
            return
        if field not in sel_list:
            sel_list.append(field)
    else:
        if field in sel_list:
            sel_list.remove(field)


def _select_all(fields, sel_list, checks):
    sel_list.clear()
    sel_list.extend(fields)
    for cb in checks.values():
        cb.value = True


def _deselect_all(sel_list, checks):
    sel_list.clear()
    for cb in checks.values():
        cb.value = False


# ---------------------------------------------------------------------------
# Порівняння
# ---------------------------------------------------------------------------

async def _do_compare(state, report_ctrl, auth_manager, results_container):
    if not state['file_bytes']:
        ui.notify('Немає файлу для порівняння', type='warning')
        return
    if not state['mapping']:
        ui.notify('Налаштуйте маппінг ключових полів', type='warning')
        return

    results_container.clear()
    state['pending_saves'].clear()
    save_btn = state.get('_save_btn')
    if save_btn:
        save_btn.set_visibility(False)

    with results_container:
        ui.spinner(size='lg').classes('mt-4')
        ui.label('Завантажуємо дані та шукаємо збіги...').classes('text-grey mt-2')

    try:
        rows = await auth_manager.execute(
            report_ctrl.compare_file_with_db,
            auth_manager.get_current_context(),
            state['file_bytes'],
            state['mapping'],
            list(state['sel_upload']),
            list(state['sel_general']),
        )

        results_container.clear()
        if not rows:
            with results_container:
                ui.label('Не знайдено жодного рядка з даними.').classes('text-warning')
            return

        state['_last_rows'] = rows
        export_btn = state.get('_export_btn')
        if export_btn:
            export_btn.set_visibility(True)

        with results_container:
            _render_compare_table(rows, state, person_ctrl=None)

    except Exception as e:
        results_container.clear()
        with results_container:
            ui.label(f'Помилка: {e}').classes('text-negative')
        ui.notify(f'Помилка порівняння: {e}', type='negative')


# ---------------------------------------------------------------------------
# Збереження pending_saves
# ---------------------------------------------------------------------------

async def _do_save(state, person_ctrl, auth_manager, save_btn):
    pending = state.get('pending_saves', {})
    if not pending:
        ui.notify('Немає змін для збереження', type='warning')
        return

    persons = list(pending.values())
    save_btn.props('loading')
    try:
        ok = await auth_manager.execute(
            person_ctrl.save_persons,
            auth_manager.get_current_context(),
            persons,
            True,   # partial_update=True — записуємо тільки заповнені поля
        )
        if ok:
            ui.notify(f'✅ Збережено {len(persons)} записів', type='positive')
            state['pending_saves'].clear()
            save_btn.set_visibility(False)
        else:
            ui.notify('Помилка збереження', type='negative')
    except Exception as e:
        ui.notify(f'Помилка: {e}', type='negative')
    finally:
        save_btn.props(remove='loading')


# ---------------------------------------------------------------------------
# Таблиця результатів
# ---------------------------------------------------------------------------

def _render_compare_table(rows: list[dict], state: dict, person_ctrl=None):
    sel_upload  = state['sel_upload']
    sel_general = state['sel_general']
    mapping     = state['mapping']
    pending     = state['pending_saves']

    found_count   = sum(1 for r in rows if r.get('found'))
    missing_count = len(rows) - found_count

    # Зворотний маппінг: upload_col → gen_field (для комбобоксів переносу)
    col_to_gen: dict[str, str] = {v: k for k, v in mapping.items()}

    # Комбобокси над стовпцями: upload_col → вибраний gen_field для переносу
    transfer_targets: dict[str, str] = {}  # {upload_col: gen_field}

    # Статистика
    with ui.row().classes('w-full gap-4 mb-2'):
        _stat_card('Рядків у файлі', len(rows),      'blue-grey')
        _stat_card('Знайдено в базі', found_count,   'positive')
        _stat_card('Не знайдено',     missing_count, 'negative')

    # -------------------------------------------------------------------
    # Комбобокси переносу — над кожним стовпцем файлу
    # -------------------------------------------------------------------
    with ui.card().classes('w-full p-3 bg-blue-50 mb-2'):
        ui.label('Перенесення даних').classes('text-caption font-bold text-blue-900 uppercase mb-2')
        ui.label(
            'Оберіть для кожного стовпця файлу відповідне поле бази — '
            'кнопка «Виправити» з\'явиться там де в базі порожньо, але у файлі є значення.'
        ).classes('text-grey text-sm mb-3')
        with ui.row().classes('w-full gap-4 flex-wrap'):
            for up_col in sel_upload:
                with ui.column().classes('gap-1').style('min-width:180px'):
                    ui.label(f'📄 {up_col}').classes('text-xs text-blue-9 font-medium')
                    default_gen = col_to_gen.get(up_col, '')
                    transfer_targets[up_col] = default_gen

                    options = ['— не переносити —'] + list(GENERAL_FIELDS)

                    def make_change(col):
                        def on_change(e):
                            val = e.value
                            transfer_targets[col] = '' if val.startswith('—') else val
                        return on_change

                    ui.select(
                        options=options,
                        value=default_gen if default_gen else '— не переносити —',
                        on_change=make_change(up_col),
                    ).props('dense outlined').style('width:100%')

    # -------------------------------------------------------------------
    # Фільтри
    # -------------------------------------------------------------------
    filters: dict[str, str] = {}

    all_columns = [
        {'name': 'row',    'label': '№',      'field': 'row',    'align': 'center',
         'style': 'width:55px', 'sortable': True},
        {'name': 'status', 'label': 'Статус', 'field': 'status', 'align': 'center',
         'style': 'width:130px', 'sortable': True},
    ]
    date_sort_fields = set()

    for col in sel_upload:
        fname = f'f_{col}'
        all_columns.append({
            'name': fname, 'label': f'📄 {col}', 'field': fname,
            'align': 'left', 'sortable': True,
        })
        # Якщо колонка схожа на дату — додаємо прихований sort_key
        if any(kw in col.lower() for kw in ('дата', 'date', 'народж', 'birth')):
            date_sort_fields.add(fname)
            sk_field = f'_sk_{fname}'
            all_columns.append({
                'name': sk_field, 'label': '', 'field': sk_field,
                'classes': 'hidden', 'headerClasses': 'hidden',
            })

    for gf in sel_general:
        fname = f'db_{gf}'
        all_columns.append({
            'name': fname, 'label': f'🗄 {gf}', 'field': fname,
            'align': 'left', 'sortable': True,
        })
        if any(kw in gf.lower() for kw in ('дата', 'date', 'народж', 'birth')):
            date_sort_fields.add(fname)
            sk_field = f'_sk_{fname}'
            all_columns.append({
                'name': sk_field, 'label': '', 'field': sk_field,
                'classes': 'hidden', 'headerClasses': 'hidden',
            })

    all_columns.append({
        'name': 'actions', 'label': 'Дія', 'field': 'actions',
        'align': 'center', 'style': 'width:100px',
    })

    # -------------------------------------------------------------------
    # Рядки таблиці
    # -------------------------------------------------------------------
    table_rows = []
    for i, r in enumerate(rows, 1):
        found = r.get('found', False)
        tr = {
            'id':           i,
            'row':          i,
            'status':       '❓ Не знайдено' if not found else '✅ Знайдено',
            'found':        found,
            'db_logical_id': r.get('db_logical_id'),   # логічний № для update_row_by_id
            'db_mil_unit':  r.get('db_mil_unit', ''),
            '_sort_order':  0 if not found else 1,
        }
        for col in sel_upload:
            val = r.get('file_data', {}).get(col, '')
            tr[f'f_{col}'] = val
            if f'f_{col}' in date_sort_fields:
                tr[f'_sk_f_{col}'] = _parse_date_for_sort(val)

        for gf in sel_general:
            val = r.get('db_data', {}).get(gf, '')
            tr[f'db_{gf}'] = val
            if f'db_{gf}' in date_sort_fields:
                tr[f'_sk_db_{gf}'] = _parse_date_for_sort(val)

        table_rows.append(tr)

    table_rows.sort(key=lambda r: r['_sort_order'])

    # -------------------------------------------------------------------
    # Рядок фільтрів
    # -------------------------------------------------------------------
    result_label = ui.label(f'Результатів: {len(table_rows)}').classes(
        'text-right text-grey w-full text-sm'
    )

    filterable = [c for c in all_columns
                  if c['name'] not in ('row', 'status', 'actions')
                  and not c['name'].startswith('_')]
    if filterable:
        with ui.row().classes('w-full gap-2 flex-wrap items-end p-2 bg-grey-1 rounded'):
            ui.label('Фільтри:').classes('text-caption text-grey self-center flex-shrink-0')
            for col_def in filterable:
                field = col_def['field']
                lbl   = col_def['label'].replace('📄 ', '').replace('🗄 ', '')

                def make_handler(f):
                    def on_change(e):
                        val = (e.value or '').strip().lower()
                        if val:
                            filters[f] = val
                        else:
                            filters.pop(f, None)
                        _apply_filters()
                    return on_change

                ui.input(
                    label=lbl, on_change=make_handler(field),
                ).props('dense outlined clearable').style('width:160px; flex-shrink:0')

    # -------------------------------------------------------------------
    # Таблиця
    # -------------------------------------------------------------------
    t = ui.table(
        columns=all_columns, rows=list(table_rows), row_key='id'
    ).classes('w-full compare-table')
    t.props('bordered separator=cell flat dense')

    t.add_slot('body-cell-status', '''
        <q-td :props="props">
            <q-badge
                :color="props.row.found ? 'positive' : 'negative'"
                :label="props.value" class="text-xs"
            />
        </q-td>
    ''')

    # Слот «Дія» — кнопка Виправити
    t.add_slot('body-cell-actions', '''
        <q-td :props="props">
            <q-btn
                v-if="props.row.found && props.row._has_fix"
                flat dense round icon="auto_fix_high" color="primary" size="sm"
                @click="() => $parent.$emit('fix_row', props.row)"
            >
                <q-tooltip>Перенести дані з файлу в базу</q-tooltip>
            </q-btn>
            <q-icon
                v-if="props.row.found && props.row._is_queued"
                name="schedule" color="orange" size="sm"
            >
                <q-tooltip>В черзі на збереження</q-tooltip>
            </q-icon>
        </q-td>
    ''')

    def _refresh_fix_flags():
        """Оновлює _has_fix для кожного рядка залежно від поточних transfer_targets."""
        for tr in t._props['rows']:
            if not tr.get('found'):
                tr['_has_fix'] = False
                continue
            has_fix = False
            for up_col in sel_upload:
                gen_field = transfer_targets.get(up_col, '')
                if not gen_field:
                    continue
                file_val = tr.get(f'f_{up_col}', '')
                db_val   = tr.get(f'db_{gen_field}', '')
                if file_val and not db_val:
                    has_fix = True
                    break
            tr['_has_fix'] = has_fix
        t.update()

    def on_fix_row(e):
        row_data     = e.args
        row_idx      = row_data.get('id')                   # порядковий номер у таблиці UI
        logical_id   = row_data.get('db_logical_id')        # логічний № (колонка A Excel)
        mil_unit     = row_data.get('db_mil_unit', '')

        if not logical_id:
            ui.notify('Не знайдено логічного № запису в базі', type='warning')
            return

        # Збираємо лише поля що переносимо (partial_update=True — решта не чіпається)
        fields_to_update = {
            COLUMN_INCREMENTAL: logical_id,   # update_row_by_id шукає саме це в колонці A
            COLUMN_MIL_UNIT:    mil_unit,
        }
        transferred = []
        for up_col in sel_upload:
            gen_field = transfer_targets.get(up_col, '')
            if not gen_field:
                continue
            file_val = row_data.get(f'f_{up_col}', '')
            db_val   = row_data.get(f'db_{gen_field}', '')
            if file_val and not db_val:
                fields_to_update[gen_field] = file_val
                transferred.append(f'{gen_field}={file_val}')

        if not transferred:
            ui.notify('Немає порожніх полів для заповнення', type='info')
            return

        try:
            person = Person.from_excel_dict(fields_to_update)
            pending[row_idx] = person

            for tr in t._props['rows']:
                if tr['id'] == row_idx:
                    tr['_is_queued'] = True
                    tr['_has_fix']   = False
                    for up_col in sel_upload:
                        gf = transfer_targets.get(up_col, '')
                        if gf and fields_to_update.get(gf):
                            tr[f'db_{gf}'] = fields_to_update[gf]
                    break
            t.update()

            save_btn = state.get('_save_btn')
            if save_btn:
                save_btn.set_visibility(True)

            ui.notify(f'Додано до збереження: {", ".join(transferred)}', type='info')

        except Exception as ex:
            ui.notify(f'Помилка підготовки: {ex}', type='negative')

    t.on('fix_row', on_fix_row)

    # Ініціалізуємо _has_fix
    for tr in t._props['rows']:
        tr['_has_fix']   = False
        tr['_is_queued'] = False
    _refresh_fix_flags()

    # Оновлюємо _has_fix при зміні transfer_targets
    ui.timer(1.5, _refresh_fix_flags)

    def _apply_filters():
        filtered = (
            list(table_rows) if not filters else [
                r for r in table_rows
                if all(fv in str(r.get(ff, '')).lower() for ff, fv in filters.items())
            ]
        )
        t._props['rows'] = filtered
        t.update()
        suffix = f'  •  фільтрів: {len(filters)}' if filters else ''
        result_label.set_text(f'Результатів: {len(filtered)} / {len(table_rows)}{suffix}')

    ui.add_head_html('''
        <style>
        .compare-table .q-table__middle {
            max-height: calc(100vh - 560px);
            min-height: 280px;
            overflow-y: auto;
        }
        .compare-table thead tr th {
            position: sticky !important;
            top: 0; z-index: 2;
            background: white;
        }
        </style>
    ''')


# ---------------------------------------------------------------------------
# Статистична картка
# ---------------------------------------------------------------------------

def _stat_card(label: str, value: int, color: str):
    with ui.card().classes(f'p-3 text-center bg-{color}-50 flex-shrink-0'):
        ui.label(str(value)).classes('text-h5 font-bold')
        ui.label(label).classes('text-xs text-grey')


# ---------------------------------------------------------------------------
# Експорт xlsx
# ---------------------------------------------------------------------------

def _export_compare(rows: list[dict], sel_upload: list, sel_general: list):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Порівняння'

    bold     = Font(bold=True)
    c_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    l_align  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    hdr_fill = PatternFill('solid', fgColor='E0E0E0')
    fill_ok  = PatternFill('solid', fgColor='C6EFCE')
    fill_err = PatternFill('solid', fgColor='FFC7CE')
    border   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )

    headers = (
        ['№', 'Статус']
        + [f'Файл: {c}' for c in sel_upload]
        + [f'База: {g}' for g in sel_general]
    )
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = bold; cell.fill = hdr_fill
        cell.alignment = c_align; cell.border = border

    for ri, r in enumerate(rows, 2):
        vals = (
            [ri - 1, 'Знайдено' if r.get('found') else 'Не знайдено']
            + [r.get('file_data', {}).get(col, '') for col in sel_upload]
            + [r.get('db_data',   {}).get(gf,  '') for gf  in sel_general]
        )
        row_fill = fill_ok if r.get('found') else fill_err
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=str(val) if val else '')
            cell.border = border
            cell.alignment = c_align if ci <= 2 else l_align
            cell.fill = row_fill

    ws.freeze_panes = 'A2'
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 14
    for i in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ui.download(buf.getvalue(), filename='compare_result.xlsx')
    ui.notify('Файл завантажується...', type='positive')
