from nicegui import ui, run

import utils.utils
from gui.controllers.person_controller import PersonController
from gui.controllers.report_controller import ReportController
from gui.services.auth_manager import AuthManager
from domain.person_filter import PersonSearchFilter
from dics.deserter_xls_dic import REVIEW_STATUS_NOT_ASSIGNED, REVIEW_STATUS_ASSIGNED, REVIEW_STATUS_CLOSED, COLUMN_INSERT_DATE
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from datetime import datetime

def render_place_report_page(report_ctrl:ReportController, person_ctrl: PersonController, auth_manager: AuthManager):
    state = {
        'place_rows': [],
        'unit_rows': [],
        'columns': [],
        'cols_p': [],
        'cols_u': []
    }

    year_options = person_ctrl.get_column_options().get(COLUMN_INSERT_DATE, [])
    year_options = sorted([str(y) for y in year_options], reverse=True)

    with ui.column().classes('w-full items-center p-4'):
        ui.label('Аналітика: Загальний стан справ').classes('text-h4 mb-6')

        with ui.row().classes('w-full max-w-4xl justify-center items-center gap-4 mb-6'):

            def on_year_change():
                if year_select.value:
                    date_from.value = None
                    date_to.value = None

            # Функція: якщо почали вводити дати -> чистимо рік
            def on_date_change():
                if date_from.value or date_to.value:
                    year_select.value = None

            # Додамо фільтр по роках, якщо треба (опціонально)
            year_select = ui.select(year_options, label='Рік СЗЧ', on_change=on_year_change).classes('w-32')

            date_from = ui.input('З дати', on_change=on_date_change).props('type=date').classes('w-40')
            date_to = ui.input('По дату', on_change=on_date_change).props('type=date').classes('w-40')

            generate_btn = ui.button('Сформувати', icon='analytics',
                                     on_click=lambda: do_report()) \
                .props('elevated color="primary"')


        ui.label('Аналіз за місцем залишення (Обставини)').classes('text-h6 mt-4')
        ui.button('Експорт (Обставини)', icon='download', color='green',
                  on_click=lambda: export_place_report_to_excel(state['place_rows'], state['cols_p'], "Обставини")) \
            .props('flat dense').bind_visibility_from(state, 'place_rows', backward=lambda r: len(r) > 0)
        place_container = ui.column().classes('w-full items-center mt-2')

        ui.label('Аналіз за підрозділами').classes('text-h6 mt-8')
        ui.button('Експорт (Підрозділи)', icon='download', color='green',
                  on_click=lambda: export_place_report_to_excel(state['unit_rows'], state['cols_u'], "Підрозділи")) \
            .props('flat dense').bind_visibility_from(state, 'unit_rows', backward=lambda r: len(r) > 0)
        unit_container = ui.column().classes('w-full items-center mt-2')


    async def do_report():
        place_container.clear()
        unit_container.clear()

        with place_container:
            loading = ui.spinner(size='xl')

        filt = PersonSearchFilter()
        filt.des_date_from = date_from.value
        filt.des_date_to = date_to.value

        try:
            # Створюємо фільтр (можна розширити)
            filt = PersonSearchFilter()
            if year_select.value:
                filt.des_year = [year_select.value]

            filt.des_date_from = date_from.value
            filt.des_date_to = date_to.value


            # Викликаємо твій новий метод з контролера
            data = await auth_manager.execute(report_ctrl.get_general_state_report, auth_manager.get_current_context(), filt)

            place_container.clear()

            if not data:
                ui.notify('Дані за вказаними критеріями відсутні', type='warning')
                return

            with place_container:
                rows_p, cols_p = build_report_table(data.get('places', {}), 'Місце СЗЧ', show_dynamic_sources=False)
                state['place_rows'] = rows_p
                state['columns'] = cols_p
                state['cols_p'] = cols_p

            with unit_container:
                rows_u, cols_u = build_report_table(data.get('units', {}), 'Підрозділ', show_dynamic_sources=True)
                state['unit_rows'] = rows_u
                # Оновлюємо колонки, бо в підрозділах їх більше
                state['columns'] = cols_u
                state['cols_u'] = cols_u

        except Exception as e:
            ui.notify(f'Помилка: {e}', type='negative')


PLACE_ORDER = ['рвбз', 'нц', 'ппд', 'лікування', 'відпустка', 'відрядження']

def sort_by_priority(item_name):
    """Повертає індекс для сортування: спочатку за списком, потім за алфавітом."""
    name_lower = str(item_name).lower().strip()
    if name_lower in PLACE_ORDER:
        return (0, PLACE_ORDER.index(name_lower))
    return (1, name_lower)


def build_report_table(data, first_col_label, show_dynamic_sources=False):
    rows = []

    # 1. Визначаємо динамічні джерела (тільки якщо потрібно)
    sorted_sources = []
    if show_dynamic_sources:
        all_sources = set()
        for stats in data.values():
            if 'dynamic_places' in stats:
                all_sources.update(stats['dynamic_places'].keys())
        sorted_sources = sorted(list(all_sources), key=sort_by_priority)

    # 2. Шаблон для "РАЗОМ"
    grand_total = {
        'place': 'РАЗОМ ПО ВЧ',
        REVIEW_STATUS_NOT_ASSIGNED: 0,
        REVIEW_STATUS_ASSIGNED: 0,
        REVIEW_STATUS_CLOSED: 0,
        'term_under_10': 0,
        'term_10_30': 0,
        'term_over_30': 0,
        'total': 0,
        'is_grand_total': True
    }
    if show_dynamic_sources:
        for src in sorted_sources:
            grand_total[f'src_{src}'] = 0

    # 3. Формуємо ряди
    for place, stats in data.items():
        row = {'place': place, 'is_grand_total': False, **stats}

        if show_dynamic_sources:
            dynamic = stats.get('dynamic_places', {})
            for src in sorted_sources:
                val = dynamic.get(src, 0)
                row[f'src_{src}'] = val
                grand_total[f'src_{src}'] += val

        rows.append(row)

        for key in grand_total:
            if key not in ['place', 'is_grand_total'] and not key.startswith('src_'):
                grand_total[key] += stats.get(key, 0)

    # --- ЛОГІКА СОРТУВАННЯ РЯДКІВ ---
    if not show_dynamic_sources:
        # ПЕРША ТАБЛИЦЯ: за вашим списком пріоритетів
        rows.sort(key=lambda x: sort_by_priority(x['place']))
    else:
        # ДРУГА ТАБЛИЦЯ (Підрозділи): за алфавітом
        rows.sort(key=lambda x: str(x['place']).lower())

    rows.append(grand_total)

    # 4. Колонки
    columns = [{'name': 'place', 'label': first_col_label, 'field': 'place', 'align': 'left'}]

    if show_dynamic_sources:
        for src in sorted_sources:
            columns.append({
                'name': f'src_{src}',
                'label': src,
                'field': f'src_{src}',
                'headerClasses': 'bg-green-50'
            })

        # НОВИЙ ПОРЯДОК СТАНДАРТНИХ КОЛОНОК
    columns.extend([
        {'name': 'u10', 'label': 'до 10 діб', 'field': 'term_under_10', 'headerClasses': 'bg-orange-50'},
        {'name': 'u30', 'label': '10-30 діб', 'field': 'term_10_30', 'headerClasses': 'bg-orange-50'},
        {'name': 'o30', 'label': '> 30 діб', 'field': 'term_over_30', 'headerClasses': 'bg-red-50'},
        {'name': 'assigned', 'label': 'Всього призначено', 'field': REVIEW_STATUS_ASSIGNED, 'headerClasses': 'bg-orange-100 font-bold'},
        {'name': 'not_assigned', 'label': 'Не призначено', 'field': REVIEW_STATUS_NOT_ASSIGNED, 'headerClasses': 'bg-amber-100'},
        {'name': 'closed', 'label': 'Закрито', 'field': REVIEW_STATUS_CLOSED, 'headerClasses': 'bg-green-100'},
        {'name': 'total', 'label': 'Всього', 'field': 'total', 'headerClasses': 'bg-blue-100 font-bold'},
    ])

    # 5. Рендеринг з адаптивною шапкою
    table = ui.table(columns=columns, rows=rows, row_key='place').classes('w-full max-w-full overflow-x-auto')
    table.props('bordered separator=cell flat dense')

    src_colspan = len(sorted_sources)
    # Якщо динамічних стовпців немає, ця секція просто не відобразиться коректно,
    # тому робимо умову в f-рядку для colspan:
    sources_header = f'<q-th colspan="{src_colspan}" class="bg-green-1 text-bold">Обставини (звідки СЗЧ)</q-th>' if show_dynamic_sources else ''

    table.add_slot('header', f'''
        <q-tr>
            <q-th colspan="1" class="bg-grey-2 text-bold">Найменування</q-th>
            {sources_header}
            <q-th colspan="3" class="bg-indigo-1 text-bold">Статус розслідування</q-th>
            <q-th colspan="3" class="bg-orange-1 text-bold">Тривалість СЗЧ (призначено)</q-th>
            <q-th colspan="1" class="bg-blue-2 text-bold">Загалом</q-th>
        </q-tr>
        <q-tr>
            <q-th v-for="col in props.cols" :key="col.name" :props="props">
                {{{{ col.label }}}}
            </q-th>
        </q-tr>
    ''')

    return rows, columns


def export_place_report_to_excel(rows, columns, sheet_name_suffix=""):
    if not rows:
        ui.notify('Немає даних для експорту', type='warning')
        return

    wb = Workbook()
    ws = wb.active
    ws.title = f"{sheet_name_suffix}"

    # --- ВИЗНАЧЕННЯ КОЛЬОРІВ (HEX) ---
    fill_cream = PatternFill("solid", fgColor="FFF9C4")  # Кремовий (терміни + призначено)
    fill_orange = PatternFill("solid", fgColor="FFE0B2")  # Лайтово-померанчевий (не призначено)
    fill_green = PatternFill("solid", fgColor="C8E6C9")  # Блідо-зелений (закрито)
    fill_total = PatternFill("solid", fgColor="BBDEFB")  # Блакитний (всього)
    header_fill = PatternFill("solid", fgColor="EEEEEE")

    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    vertical_align = Alignment(horizontal="center", vertical="center", textRotation=90)

    # Визначаємо індекс, де починаються колонки термінів (після "звідки")
    # У нашому новому порядку перша колонка термінів — 'u10'
    idx_terms_start = next((i for i, c in enumerate(columns) if c['name'] == 'u10'), len(columns))

    # --- ШАПКА (Рядок 2) ---
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=idx, value=col['label'])
        cell.font = bold_font
        cell.border = thin_border
        cell.fill = header_fill

        # Вертикальний текст для "Звідки" (колонки між першою та термінами)
        if idx > 1:
            cell.alignment = vertical_align
            ws.column_dimensions[cell.column_letter].width = 5
        else:
            cell.alignment = center_align
            ws.column_dimensions[cell.column_letter].width = 15

    ws.column_dimensions['A'].width = 35

    # --- ДАНІ ТА ФАРБУВАННЯ (Рядок 3+) ---
    for r_idx, row_data in enumerate(rows, start=3):
        is_total_row = row_data.get('is_grand_total', False)

        for c_idx, col_def in enumerate(columns, start=1):
            val = row_data.get(col_def['field'])
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            cell.alignment = center_align if c_idx > 1 else Alignment(horizontal="left")

            # Логіка кольорів для стовпчиків:
            if c_idx > 1:  # не фарбуємо першу колонку назв, крім підсумку
                c_name = col_def['name']

                # Перші 4 після "звідки" (до 10, до 30, >30, призначено) -> Кремовий
                if c_name in ['u10', 'u30', 'o30', 'assigned']:
                    cell.fill = fill_cream

                # Не призначено -> Лайтово-померанчевий
                elif c_name == 'not_assigned':
                    cell.fill = fill_orange

                # Закрито -> Блідо-зелений
                elif c_name == 'closed':
                    cell.fill = fill_green

                # Всього (остання) -> Блакитний
                elif c_name == 'total':
                    cell.fill = fill_total
                    cell.font = bold_font

            if is_total_row:
                cell.font = bold_font
                # Для підсумкового рядка можна змінити стиль,
                # або залишити колір колонки, але додати жирний шрифт

    ws.freeze_panes = 'A3'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"General_Report_{sheet_name_suffix}.xlsx"
    ui.download(buffer.getvalue(), filename=filename)