from nicegui import ui, run
from dics.deserter_xls_dic import *
from gui.services.request_context import RequestContext

import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

last_query = {}

def search_page(report_ctrl, person_ctrl, ctx: RequestContext):
    global last_query
    state = {'rows': [], 'columns': []}

    year_options = person_ctrl.get_column_options().get(COLUMN_INSERT_DATE, [])

    with ui.column().classes('w-full items-center p-4'):
        ui.label('Звіт Додаток №2, по підрозділам').classes('text-h4 mb-4')

        with ui.row().classes('w-full max-w-4xl items-center gap-4'):
            year_filter = ui.select(
                options=year_options,
                label=COLUMN_INSERT_DATE,
                clearable = True
            ).classes('w-64').props('use-chips stack-label').on('update:model-value', lambda: do_report())

            # Додаємо пошук по Enter
            search_btn = ui.button('Пошук', icon='search', on_click=lambda: do_report()).props('elevated')

            export_btn = export_btn =ui.button('Експорт', icon='download', color='green',
                                   on_click=lambda: export_to_excel(state['rows'], state['columns']))

        results_container = ui.column().classes('w-full items-center mt-6')
        export_btn.bind_visibility_from(results_container, 'visible')

    async def do_report():
        global last_query
        selected_year = year_filter.value
        year_val = None if selected_year == 'Всі роки' or not selected_year else selected_year

        if not year_val:
            ui.notify('Введіть рік для репорту', type='warning')
            return

        last_query['year'] = year_val

        year_filter.disable()
        search_btn.disable()
        export_btn.disable()

        with results_container:
            ui.spinner(size='lg').classes('mt-10')
            ui.label('Компайлінг звіту...').classes('text-grey')

        try:
            data = await run.io_bound(report_ctrl.do_subunit_desertion_report,ctx, [year_val])

            # 3. Очищуємо спіннер після отримання даних
            results_container.clear()

            if not data:
                ui.notify('Нічого не скомпілілося. А якого?', type='negative')
                return

            # 4. Малюємо таблицю з результатами
            with results_container:
                rows, columns = results_ui(data, report_ctrl)
                state['rows'] = rows
                state['columns'] = columns

        except Exception as e:
            results_container.clear()
            ui.notify(f'Помилка пошуку: {e}', type='negative')
        finally:
            year_filter.enable()
            search_btn.enable()
            export_btn.enable()

def results_ui(data, report_ctrl):
    if not data:
        ui.label('Дані відсутні').classes('text-grey')
        return

    rows = []

    # 1. Створюємо об'єкт для глобального підсумку (Разом по всій частині)
    grand_total = {
        'main': 'РАЗОМ ПО ЧАСТИНІ',
        'sub': '',
        's_under': 0, 'o_under': 0, 's_over': 0, 'o_over': 0,
        's_ret_mu': 0, 's_ret_res': 0, 'o_ret': 0,
        's_diff': 0, 'o_diff': 0,
        'expl_not_ass': 0, 'expl_not_exe': 0, 'expl_not_cls': 0,
        's_dupl': 0, 'o_dupl': 0,

        'un_sold_des': 0,
        'un_sold_ret': 0,
        'un_serg_des': 0,
        'un_serg_ret': 0,
        'un_offc_des': 0,
        'un_offc_ret': 0,
        'st_term': 0,
        'st_call': 0,
        'st_contr': 0,
        'pl_ppd': 0,
        'pl_rvbz': 0,
        'pl_other': 0,

        'weapon': 0,

        'rev_specified_total': 0,
        'rev_specified_of': 0,
        'rev_dbr_notif_total': 0,
        'rev_dbr_notif_of': 0,
        'rev_dbr_mater_total': 0,
        'rev_dbr_mater_of': 0,
        'rev_dbr_nonerdr_total': 0,
        'rev_dbr_nonerdr_of': 0,
        'rev_dbr_erdr_total': 0,
        'rev_dbr_erdr_of': 0,
        'rev_suspend_total': 0,
        'rev_suspend_of': 0,
        'rev_courts_total': 0,
        'rev_courts_of': 0,
        'rev_punish_total': 0,
        'rev_punish_of': 0,
        'rev_nonevil_total': 0,
        'rev_nonevil_of': 0,

        'is_grand_total': True  # Спеціальний прапорець для стилізації
    }

    # Сортуємо ключі для стабільного виведення
    for main_unit in sorted(data.keys()):
        sub_units_data = data[main_unit]

        # Спочатку рахуємо суму для всього основного підрозділу
        unit_total = {
            'main': main_unit,
            'sub': 'ВСЬОГО',  # Мітка для ідентифікації підсумкового рядка
            's_under': 0, 'o_under': 0, 's_over': 0, 'o_over': 0,
            's_ret_mu': 0, 's_ret_res': 0, 'o_ret': 0,
            's_diff': 0, 'o_diff': 0,
            'expl_not_ass': 0, 'expl_not_exe': 0, 'expl_not_cls': 0,
            's_dupl': 0, 'o_dupl': 0,

            'un_sold_des': 0,
            'un_sold_ret': 0,
            'un_serg_des': 0,
            'un_serg_ret': 0,
            'un_offc_des': 0,
            'un_offc_ret': 0,
            'st_term': 0,
            'st_call': 0,
            'st_contr': 0,
            'pl_ppd': 0,
            'pl_rvbz': 0,
            'pl_other': 0,

            'weapon': 0,

            'rev_specified_total': 0,
            'rev_specified_of': 0,
            'rev_dbr_notif_total': 0,
            'rev_dbr_notif_of': 0,
            'rev_dbr_mater_total': 0,
            'rev_dbr_mater_of': 0,
            'rev_dbr_nonerdr_total': 0,
            'rev_dbr_nonerdr_of': 0,
            'rev_dbr_erdr_total': 0,
            'rev_dbr_erdr_of': 0,
            'rev_suspend_total': 0,
            'rev_suspend_of': 0,
            'rev_courts_total': 0,
            'rev_courts_of': 0,
            'rev_punish_total': 0,
            'rev_punish_of': 0,
            'rev_nonevil_total': 0,
            'rev_nonevil_of': 0,

            'is_total': True  # Прапорець для стилізації
        }

        temp_sub_rows = []
        for sub_unit, stats in sub_units_data.items():
            s_under = stats['рядовий_сержант']['under_3'] + stats['офіцер']['under_3'] # all
            s_over = stats['рядовий_сержант']['over_3'] + stats['офіцер']['over_3'] # all
            o_under = stats['офіцер']['under_3']
            o_over = stats['офіцер']['over_3']
            s_ret_mu = stats['рядовий_сержант']['ret_mu']
            s_ret_res = stats['рядовий_сержант']['ret_res']
            o_ret_mu = stats['офіцер']['ret_mu']
            s_dupl = stats['рядовий']['dupl'] + stats['сержант']['dupl'] + stats['офіцер']['dupl'] # документація каже, що треба всіх сюди
            o_dupl = stats['офіцер']['dupl']

            un_sold_des = stats['рядовий']['un_des'] if stats['рядовий']['un_des'] else 0
            un_sold_ret = stats['рядовий']['un_ret'] if stats['рядовий']['un_ret'] else 0
            un_serg_des = stats['сержант']['un_des'] if stats['сержант']['un_des'] else 0
            un_serg_ret = stats['сержант']['un_ret'] if stats['сержант']['un_ret'] else 0
            un_offc_des = stats['офіцер']['un_des'] if stats['офіцер']['un_des'] else 0
            un_offc_ret = stats['офіцер']['un_ret'] if stats['офіцер']['un_ret'] else 0
            st_term = stats['all']['st_term'] if stats['all']['st_term'] else 0
            st_call = stats['all']['st_call'] if stats['all']['st_call'] else 0
            st_contr = stats['all']['st_contr'] if stats['all']['st_contr'] else 0
            pl_ppd = stats['all']['pl_ppd'] if stats['all']['pl_ppd'] else 0
            pl_rvbz = stats['all']['pl_rvbz'] if stats['all']['pl_rvbz'] else 0
            pl_other = stats['all']['pl_other'] if stats['all']['pl_other'] else 0

            expl_not_ass = stats['all'][REVIEW_STATUS_NOT_ASSIGNED]
            expl_not_exe = stats['all'][REVIEW_STATUS_EXECUTING]
            expl_not_cls = stats['all'][REVIEW_STATUS_CLOSED]

            # 2. Допоміжна функція або логіка для підрахунку (Total = Сержант + Офіцер, Of = Офіцер)
            weapon = stats['all']['weapon']

            def get_rev_stat(key):
                s_val = stats['all'].get(key, 0)
                o_val = stats['офіцер'].get(key, 0)
                return (s_val + o_val), o_val

            # 3. Розрахунок нових полів
            rev_specified_total, rev_specified_of = get_rev_stat('rev_specified')
            rev_dbr_notif_total, rev_dbr_notif_of = get_rev_stat('rev_dbr_notif')
            rev_dbr_mater_total, rev_dbr_mater_of = get_rev_stat('rev_dbr_mater')
            rev_dbr_nonerdr_total, rev_dbr_nonerdr_of = get_rev_stat('rev_dbr_nonerdr')
            rev_dbr_erdr_total, rev_dbr_erdr_of = get_rev_stat('rev_dbr_erdr')
            rev_suspend_total, rev_suspend_of = get_rev_stat('rev_suspend')
            rev_courts_total, rev_courts_of = get_rev_stat('rev_courts')
            rev_punish_total, rev_punish_of = get_rev_stat('rev_punish')
            rev_nonevil_total, rev_nonevil_of = get_rev_stat('rev_nonevil')

            # Додаємо до підсумку підрозділу
            unit_total['s_under'] += s_under
            unit_total['s_over'] += s_over
            unit_total['o_under'] += o_under
            unit_total['o_over'] += o_over
            unit_total['s_ret_mu'] += s_ret_mu
            unit_total['s_ret_res'] += s_ret_res
            unit_total['o_ret'] += o_ret_mu
            unit_total['s_diff'] = unit_total['s_diff'] + s_under + s_over - s_ret_mu - s_ret_res
            unit_total['o_diff'] = unit_total['o_diff'] + o_under + o_over - o_ret_mu
            unit_total['expl_not_ass'] += expl_not_ass
            unit_total['expl_not_exe'] += expl_not_exe
            unit_total['expl_not_cls'] += expl_not_cls
            unit_total['s_dupl'] += s_dupl
            unit_total['o_dupl'] += o_dupl

            unit_total['un_sold_des'] += un_sold_des
            unit_total['un_sold_ret'] += un_sold_ret
            unit_total['un_serg_des'] += un_serg_des
            unit_total['un_serg_ret'] += un_serg_ret
            unit_total['un_offc_des'] += un_offc_des
            unit_total['un_offc_ret'] += un_offc_ret
            unit_total['st_term'] += st_term
            unit_total['st_call'] += st_call
            unit_total['st_contr'] += st_contr
            unit_total['pl_ppd'] += pl_ppd
            unit_total['pl_rvbz'] += pl_rvbz
            unit_total['pl_other'] += pl_other

            # 4. Додаємо до підсумку підрозділу (unit_total)
            unit_total['weapon'] += weapon

            unit_total['rev_specified_total'] += rev_specified_total
            unit_total['rev_specified_of'] += rev_specified_of
            unit_total['rev_dbr_notif_total'] += rev_dbr_notif_total
            unit_total['rev_dbr_notif_of'] += rev_dbr_notif_of
            unit_total['rev_dbr_mater_total'] += rev_dbr_mater_total
            unit_total['rev_dbr_mater_of'] += rev_dbr_mater_of
            unit_total['rev_dbr_nonerdr_total'] += rev_dbr_nonerdr_total
            unit_total['rev_dbr_nonerdr_of'] += rev_dbr_nonerdr_of
            unit_total['rev_dbr_erdr_total'] += rev_dbr_erdr_total
            unit_total['rev_dbr_erdr_of'] += rev_dbr_erdr_of
            unit_total['rev_suspend_total'] += rev_suspend_total
            unit_total['rev_suspend_of'] += rev_suspend_of
            unit_total['rev_courts_total'] += rev_courts_total
            unit_total['rev_courts_of'] += rev_courts_of
            unit_total['rev_punish_total'] += rev_punish_total
            unit_total['rev_punish_of'] += rev_punish_of
            unit_total['rev_nonevil_total'] += rev_nonevil_total
            unit_total['rev_nonevil_of'] += rev_nonevil_of

            # Додаємо до ГЛОБАЛЬНОГО підсумку
            grand_total['s_under'] += s_under
            grand_total['s_over'] += s_over
            grand_total['o_under'] += o_under
            grand_total['o_over'] += o_over
            grand_total['s_ret_mu'] += s_ret_mu
            grand_total['s_ret_res'] += s_ret_res
            grand_total['o_ret'] += o_ret_mu
            grand_total['s_diff'] = grand_total['s_diff'] + s_under + s_over - s_ret_mu - s_ret_res
            grand_total['o_diff'] = grand_total['o_diff'] + o_under + o_over - o_ret_mu
            grand_total['expl_not_ass'] += expl_not_ass
            grand_total['expl_not_exe'] += expl_not_exe
            grand_total['expl_not_cls'] += expl_not_cls
            grand_total['s_dupl'] += s_dupl
            grand_total['o_dupl'] += o_dupl
            grand_total['un_sold_des'] += un_sold_des
            grand_total['un_sold_ret'] += un_sold_ret
            grand_total['un_serg_des'] += un_serg_des
            grand_total['un_serg_ret'] += un_serg_ret
            grand_total['un_offc_des'] += un_offc_des
            grand_total['un_offc_ret'] += un_offc_ret
            grand_total['st_term'] += st_term
            grand_total['st_call'] += st_call
            grand_total['st_contr'] += st_contr
            grand_total['pl_ppd'] += pl_ppd
            grand_total['pl_rvbz'] += pl_rvbz
            grand_total['pl_other'] += pl_other
            grand_total['weapon'] += weapon
            grand_total['rev_specified_total'] += rev_specified_total
            grand_total['rev_specified_of'] += rev_specified_of
            grand_total['rev_dbr_notif_total'] += rev_dbr_notif_total
            grand_total['rev_dbr_notif_of'] += rev_dbr_notif_of
            grand_total['rev_dbr_mater_total'] += rev_dbr_mater_total
            grand_total['rev_dbr_mater_of'] += rev_dbr_mater_of
            grand_total['rev_dbr_nonerdr_total'] += rev_dbr_nonerdr_total
            grand_total['rev_dbr_nonerdr_of'] += rev_dbr_nonerdr_of
            grand_total['rev_dbr_erdr_total'] += rev_dbr_erdr_total
            grand_total['rev_dbr_erdr_of'] += rev_dbr_erdr_of
            grand_total['rev_suspend_total'] += rev_suspend_total
            grand_total['rev_suspend_of'] += rev_suspend_of
            grand_total['rev_courts_total'] += rev_courts_total
            grand_total['rev_courts_of'] += rev_courts_of
            grand_total['rev_punish_total'] += rev_punish_total
            grand_total['rev_punish_of'] += rev_punish_of
            grand_total['rev_nonevil_total'] += rev_nonevil_total
            grand_total['rev_nonevil_of'] += rev_nonevil_of

            temp_sub_rows.append({
                'main': '',
                'sub': sub_unit,
                's_under': s_under,
                'o_under': o_under,
                's_over': s_over,
                'o_over': o_over,
                's_ret_mu': s_ret_mu,
                's_ret_res': s_ret_res,
                'o_ret': o_ret_mu,
                's_diff': s_under + s_over - s_ret_mu - s_ret_res,
                'o_diff': o_under + o_over - o_ret_mu,
                'expl_not_ass': expl_not_ass,
                'expl_not_exe': expl_not_exe,
                'expl_not_cls': expl_not_cls,
                's_dupl': s_dupl,
                'o_dupl': o_dupl,
                'un_sold_des': un_sold_des,
                'un_sold_ret': un_sold_ret,
                'un_serg_des': un_serg_des,
                'un_serg_ret': un_serg_ret,
                'un_offc_des': un_offc_des,
                'un_offc_ret': un_offc_ret,
                'st_term': st_term,
                'st_call': st_call,
                'st_contr': st_contr,
                'pl_ppd': pl_ppd,
                'pl_rvbz': pl_rvbz,
                'pl_other': pl_other,

                'weapon': weapon,
                'rev_specified_total': rev_specified_total,
                'rev_specified_of': rev_specified_of,
                'rev_dbr_notif_total': rev_dbr_notif_total,
                'rev_dbr_notif_of': rev_dbr_notif_of,
                'rev_dbr_mater_total': rev_dbr_mater_total,
                'rev_dbr_mater_of': rev_dbr_mater_of,
                'rev_dbr_nonerdr_total': rev_dbr_nonerdr_total,
                'rev_dbr_nonerdr_of': rev_dbr_nonerdr_of,
                'rev_dbr_erdr_total': rev_dbr_erdr_total,
                'rev_dbr_erdr_of': rev_dbr_erdr_of,
                'rev_suspend_total': rev_suspend_total,
                'rev_suspend_of': rev_suspend_of,
                'rev_courts_total': rev_courts_total,
                'rev_courts_of': rev_courts_of,
                'rev_punish_total': rev_punish_total,
                'rev_punish_of': rev_punish_of,
                'rev_nonevil_total': rev_nonevil_total,
                'rev_nonevil_of': rev_nonevil_of,

                'is_total': False
            })

        # Рахуємо фінальні суми для підсумкового рядка
        unit_total['s_total'] = unit_total['s_under'] + unit_total['s_over']
        unit_total['o_total'] = unit_total['o_under'] + unit_total['o_over']
        unit_total['s_total_ret'] = unit_total['s_ret_mu'] + unit_total['s_ret_res']

        # Додаємо в загальний список: спочатку підсумок підрозділу, потім деталізацію
        # Додаємо підрозділ у загальний список
        rows.append(unit_total)
        temp_sub_rows.sort(key=lambda x: x['sub'])

        for r in temp_sub_rows:
            r['s_total'] = r['s_under'] + r['s_over']
            r['o_total'] = r['o_under'] + r['o_over']
            r['s_total_ret'] = r['s_ret_mu'] + r['s_ret_res']
            rows.append(r)
        # rows.extend(temp_sub_rows)

    # 3. Рахуємо фінальні суми для ГЛОБАЛЬНОГО підсумку
    grand_total['s_total'] = grand_total['s_under'] + grand_total['s_over']
    grand_total['o_total'] = grand_total['o_under'] + grand_total['o_over']
    grand_total['s_total_ret'] = grand_total['s_ret_mu'] + grand_total['s_ret_res']

    # ВСТАВЛЯЄМО ГЛОБАЛЬНИЙ ПІДСУМОК НА ПОЧАТОК
    rows.insert(0, grand_total)

    # Визначення колонок (залишається майже таким самим)
    columns = [
        {'name': 'main', 'label': 'Підрозділ', 'field': 'main', 'align': 'left'},
        {'name': 'sub', 'label': 'Саб-підрозділ', 'field': 'sub', 'align': 'left'},
        {'name': 's_under', 'label': '< 3 (Всього)', 'field': 's_under'},
        {'name': 'o_under', 'label': '< 3 (Офіц.)', 'field': 'o_under'},
        {'name': 's_over', 'label': '>= 3 (Всього)', 'field': 's_over'},
        {'name': 'o_over', 'label': '>= 3 (Офіц.)', 'field': 'o_over'},

        {'name': 's_total', 'label': 'Всього', 'field': 's_total', 'headerClasses': 'bg-blue-100'},
        {'name': 'o_total', 'label': 'Всього (Офіц.)', 'field': 'o_total', 'headerClasses': 'bg-blue-100'},
        {'name': 's_ret_mu', 'label': 'У В/Ч (С/С)', 'field': 's_ret_mu', 'headerClasses': 'bg-unit-return'},
        {'name': 's_ret_res', 'label': 'В БРез (С/С)', 'field': 's_ret_res', 'headerClasses': 'bg-unit-return'},
        {'name': 's_total_ret', 'label': 'Всього (С/С)', 'field': 's_total_ret', 'headerClasses': 'bg-blue-100'},
        {'name': 'o_ret', 'label': 'В т.ч. Офіц.', 'field': 'o_ret', 'headerClasses': 'bg-unit-return'},

        {'name': 's_diff', 'label': 'Кільість СЗЧ (С/С)', 'field': 's_diff', 'headerClasses': 'bg-blue-100'},
        {'name': 'o_diff', 'label': 'Кількість СЗЧ (Офіц.)', 'field': 'o_diff', 'headerClasses': 'bg-blue-100'},

        {'name': 'expl_not_ass', 'label': 'Не призначено', 'field': 'expl_not_ass'},
        {'name': 'expl_not_exe', 'label': 'Проводяться', 'field': 'expl_not_exe'},
        {'name': 'expl_not_cls', 'label': 'Завершено', 'field': 'expl_not_cls'},

        {'name': 's_dupl', 'label': 'Всього', 'field': 's_dupl'},
        {'name': 'o_dupl', 'label': 'Офіцерів', 'field': 'o_dupl'},

        {'name': 'un_sold_des', 'label': 'СЗЧ Солдат', 'field': 'un_sold_des'},
        {'name': 'un_sold_ret', 'label': 'Пов. Солдат', 'field': 'un_sold_ret', 'headerClasses': 'bg-unit-return'},
        {'name': 'un_serg_des', 'label': 'СЗЧ Сержант', 'field': 'un_serg_des'},
        {'name': 'un_serg_ret', 'label': 'Пов. Сержант', 'field': 'un_serg_ret', 'headerClasses': 'bg-unit-return'},
        {'name': 'un_offc_des', 'label': 'СЗЧ Офіцер', 'field': 'un_offc_des'},
        {'name': 'un_offc_ret', 'label': 'Пов. Офіцер', 'field': 'un_offc_ret', 'headerClasses': 'bg-unit-return'},

        {'name': 'st_term', 'label': 'Строкова служба', 'field': 'st_term'},
        {'name': 'st_call', 'label': 'За призивом', 'field': 'st_call'},
        {'name': 'st_contr', 'label': 'За контрактом', 'field': 'st_contr'},

        {'name': 'pl_ppd', 'label': 'З ППД', 'field': 'pl_ppd'},
        {'name': 'pl_rvbz', 'label': 'З РВБЗ', 'field': 'pl_rvbz'},
        {'name': 'pl_other', 'label': 'Інше', 'field': 'pl_other'},

        {'name': 'weapon', 'label': '', 'field': 'weapon'},

        {'name': 'rev_specified_total', 'label': 'Всього', 'field': 'rev_specified_total'},
        {'name': 'rev_specified_of', 'label': 'В т.ч офіцерів', 'field': 'rev_specified_of'},
        {'name': 'rev_dbr_notif_total', 'label': 'Всього', 'field': 'rev_dbr_notif_total'},
        {'name': 'rev_dbr_notif_of', 'label': 'В т.ч офіцерів', 'field': 'rev_dbr_notif_of'},
        {'name': 'rev_dbr_mater_total', 'label': 'Всього', 'field': 'rev_dbr_mater_total'},
        {'name': 'rev_dbr_mater_of', 'label': 'В т.ч офіцерів', 'field': 'rev_dbr_mater_of'},
        {'name': 'rev_dbr_nonerdr_total', 'label': 'Всього', 'field': 'rev_dbr_nonerdr_total'},
        {'name': 'rev_dbr_nonerdr_of', 'label': 'В т.ч офіцерів', 'field': 'rev_dbr_nonerdr_of'},
        {'name': 'rev_dbr_erdr_total', 'label': 'Всього', 'field': 'rev_dbr_erdr_total'},
        {'name': 'rev_dbr_erdr_of', 'label': 'В т.ч офіцерів', 'field': 'rev_dbr_erdr_of'},
        {'name': 'rev_suspend_total', 'label': 'Всього', 'field': 'rev_suspend_total'},
        {'name': 'rev_suspend_of', 'label': 'В т.ч офіцерів', 'field': 'rev_suspend_of'},
        {'name': 'rev_courts_total', 'label': 'Всього', 'field': 'rev_courts_total'},
        {'name': 'rev_courts_of', 'label': 'В т.ч офіцерів', 'field': 'rev_courts_of'},
        {'name': 'rev_punish_total', 'label': 'Всього', 'field': 'rev_punish_total'},
        {'name': 'rev_punish_of', 'label': 'В т.ч офіцерів', 'field': 'rev_punish_of'},
        {'name': 'rev_nonevil_total', 'label': 'Всього', 'field': 'rev_nonevil_total'},
        {'name': 'rev_nonevil_of', 'label': 'В т.ч офіцерів', 'field': 'rev_nonevil_of'},
    ]

    table = ui.table(columns=columns, rows=rows, row_key='index').classes('w-full report-table')
    table.props('bordered separator=cell flat')

    # Додаємо складний хедер (як у минулому кроці)
    table.add_slot('header', '''
            <q-tr>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Підрозділи</q-th>
                <q-th colspan="6" class="text-center bg-grey-3 text-bold text-subtitle1">Випадки СЗЧ</q-th>
                <q-th colspan="4" class="text-center bg-grey-3 text-bold text-subtitle1">Повернення</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Особового складу в СЗЧ</q-th>
                <q-th colspan="3" class="text-center bg-grey-3 text-bold text-subtitle1">Відпрацювання службових розслідувань</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">СЗЧ 2 і більше</q-th>
                <q-th colspan="6" class="text-center bg-grey-3 text-bold text-subtitle1">Унікальні випадки СЗЧ та повернення</q-th>
                <q-th colspan="3" class="text-center bg-grey-3 text-bold text-subtitle1">Вид служби</q-th>
                <q-th colspan="3" class="text-center bg-grey-3 text-bold text-subtitle1">Місце скоєння</q-th>
                <q-th colspan="1" class="text-center bg-grey-3 text-bold text-subtitle1">Зі зброєю</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Виведено у розпорядження</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Повідомлень про КП до ВСП, ДБР</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Матеріалів с/р до ВСП, ДБР</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Не отримано витяг з ЄРДР</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Отримано витяг з ЄРДР</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Призупинена в/служба</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Кількість вироків судів</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Відбувають покарання</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Не є суб'єктом злочину</q-th>
            </q-tr>
            <q-tr>
                <q-th v-for="col in props.cols" :key="col.name" :props="props" :class="col.header_classes">
                    {{ col.label }}
                </q-th>
            </q-tr>
        ''')

    # КАСТОМНЕ ВІДОБРАЖЕННЯ РЯДКІВ (Стилізація підсумків)
    table.add_slot('body', '''
            <q-tr :props="props" :class="props.row.is_grand_total ? 'bg-grand-total font-bold' : (props.row.is_total ? 'bg-unit-total' : '')">
                <q-td v-for="col in props.cols" :key="col.name" :props="props" 
                    :class="[
                        (col.name === 's_total' || col.name === 'o_total' || col.name === 's_total_ret' || col.name === 's_diff' || col.name === 'o_diff') ? 'bg-unit-total' : '',
                        (['s_ret_mu', 's_ret_res', 'o_ret', 'un_sold_ret', 'un_serg_ret', 'un_offc_ret'].includes(col.name)) ? 'bg-unit-return' : ''
                    ]">
                    <template v-if="col.name === 'sub' && props.row.is_total">
                        <q-badge color="primary">Підсумок</q-badge>
                    </template>
                    <template v-else-if="col.name === 'main' && props.row.is_grand_total">
                        <q-icon name="star" color="orange" /> {{ col.value }}
                    </template>
                    <template v-else>
                        {{ col.value }}
                    </template>
                </q-td>
            </q-tr>
        ''')

    return rows, columns


import xlwings as xw
from nicegui import ui


def export_to_excel(rows, columns):
    if not rows:
        ui.notify('Немає даних для експорту', type='warning')
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Додаток №2"

        # Налаштування стилів (openpyxl використовує HEX без решітки)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bold_font = Font(bold=True)

        # Кольори
        header_fill = PatternFill("solid", fgColor="EEEEEE")
        header_row2_fill = PatternFill("solid", fgColor="E0E0E0")
        grand_total_fill = PatternFill("solid", fgColor="FFE0B2")  # Помаранчевий
        unit_total_fill = PatternFill("solid", fgColor="E3F2FD")  # Блакитний (Всього)
        blue_col_fill = PatternFill("solid", fgColor="BBDEFB")  # Блакитні стовпці
        green_col_fill = PatternFill("solid", fgColor="C1F5DD")  # Зелені стовпці

        # 1. СТВОРЕННЯ ШАПКИ З МЕРДЖІНГОМ
        def set_header(range_str, text):
            ws.merge_cells(range_str)
            top_left_cell = range_str.split(':')[0]
            ws[top_left_cell].value = text
            # В openpyxl треба застосовувати стиль до кожної клітинки в об'єднаному діапазоні
            for row in ws[range_str]:
                for cell in row:
                    cell.fill = header_fill
                    cell.font = bold_font
                    cell.alignment = center_align
                    cell.border = thin_border

        set_header('A1:B1', 'Підрозділи')
        set_header('C1:H1', 'Випадки СЗЧ')
        set_header('I1:L1', 'Повернення')
        set_header('M1:N1', 'Особового складу в СЗЧ')
        set_header('O1:Q1', 'Відпрацювання службових розслідувань')
        set_header('R1:S1', 'СЗЧ 2 і більше')
        set_header('T1:Y1', 'Унікальні випадки СЗЧ та повернення')
        set_header('Z1:AB1', 'Вид служби')
        set_header('AC1:AE1', 'Місце скоєння')
        set_header('AF1:AF1', 'Зі зброєю')
        set_header('AG1:AH1', 'Виведено у розпорядження')
        set_header('AI1:AJ1', 'Повідомлень про КП до ВСП, ДБР')
        set_header('AK1:AL1', 'Матеріалів с/р до ВСП, ДБР')
        set_header('AM1:AN1', 'Не отримано витяг з ЄРДР')
        set_header('AO1:AP1', 'Отримано витяг з ЄРДР')
        set_header('AQ1:AR1', 'Призупинена в/служба')
        set_header('AS1:AT1', 'Кількість вироків судів')
        set_header('AU1:AV1', 'Відбувають покарання')
        set_header('AW1:AX1', "Не є суб'єктом злочину")

        # 2. ДРУГИЙ РЯДОК ЗАГОЛОВКІВ
        blue_cols = ['s_total', 'o_total', 's_total_ret', 's_diff', 'o_diff']

        for col_num, col in enumerate(columns, start=1):
            cell = ws.cell(row=2, column=col_num, value=col['label'])
            cell.fill = header_row2_fill
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border

        # 3. ЗАПИС ДАНИХ ТА РОЗФАРБОВКА
        for row_idx, r in enumerate(rows, start=3):
            is_grand_total = r.get('is_grand_total')
            is_total = r.get('is_total')

            for col_idx, col in enumerate(columns, start=1):
                # Безпечне отримання значення
                val = r.get(col['field'])
                if val is None:
                    val = ''
                elif not isinstance(val, (int, float)):
                    val = str(val)

                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

                # Базовий колір рядка
                if is_grand_total:
                    cell.fill = grand_total_fill
                    cell.font = bold_font
                elif is_total:
                    cell.fill = unit_total_fill
                    cell.font = bold_font
                else:
                    # Вертикальні акценти для звичайних рядків
                    if col['name'] in blue_cols:
                        cell.fill = blue_col_fill
                    elif 'bg-unit-return' in str(col.get('headerClasses', '')):
                        cell.fill = green_col_fill

        # 4. ЗАКРІПЛЕННЯ ОБЛАСТЕЙ (Freeze Panes)
        # 'C3' означає, що закріплено 2 верхні рядки та 2 ліві колонки
        ws.freeze_panes = 'C3'

        # 5. АВТОПІДБІР ШИРИНИ
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 25
        for col_idx in range(3, len(columns) + 1):
            # Перетворюємо номер колонки на літеру (3 -> 'C')
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 12

        # 6. ЗБЕРЕЖЕННЯ В ПАМ'ЯТЬ ТА ВІДПРАВКА БРАУЗЕРУ
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # ui.download відправляє файл у браузер користувача, який натиснув кнопку
        ui.download(buffer.getvalue(), filename='Додаток_2_Звіт.xlsx')
        ui.notify('Файл завантажується...', type='positive')

    except Exception as e:
        print(f"Excel Error Trace: {e}")
        import traceback
        traceback.print_exc()
        ui.notify(f'Помилка генерації файлу: {e}', type='negative')