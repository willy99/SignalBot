"""
Сторінка «ЄРДР КРАМ»
=====================
Завантаження Excel-файлу з кримінальними провадженнями КРАМ.
Автоматичне розпізнавання ПІБ / дати народження / РНОКПП з тексту.
Порівняння з основною базою СЗЧ.
Таблиця результатів з кольоровою індикацією статусу.
Кнопка «Змінити» для рядків ⚠️ — додає в чергу на запис ЄРДР в базу.
"""

import io

from nicegui import ui, events
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from dics.deserter_xls_dic import NA
from gui.services.auth_manager import AuthManager
from service.processing.processors.ErdrKramProcessor import ErdrKramRow


def render_erdr_kram_page(erdr_kram_ctrl, auth_manager: AuthManager):
    """Головна функція рендеру сторінки ЄРДР КРАМ."""

    state: dict = {
        'rows':          [],    # list[ErdrKramRow] після обробки
        'file_bytes':    None,
        'pending':       {},    # source_row -> ErdrKramRow — черга на збереження
    }

    # ------------------------------------------------------------------
    # Шапка — статична, не прокручується разом з таблицею
    # ------------------------------------------------------------------
    with ui.column().classes('w-full p-4 gap-3'):

        ui.label('ЄРДР КРАМ — звірка кримінальних проваджень').classes('text-h5 font-bold')
        ui.label(
            'Завантажте Excel-файл КРАМ (3 стовпці: опис порушення | в/ч | номер ЄРДР). '
            'Система автоматично витягне ПІБ та дату народження і знайде кожного в основній базі СЗЧ.'
        ).classes('text-grey text-sm max-w-3xl')

        # Рядок керування — upload ліворуч, кнопки праворуч
        with ui.row().classes('w-full items-center gap-4'):

            async def handle_upload(e: events.UploadEventArguments):
                state['file_bytes'] = await e.file.read()
                process_btn.enable()
                ui.notify(
                    f'Файл завантажено: {e.file.name} ({len(state["file_bytes"])} байт)',
                    type='info',
                )

            ui.upload(
                label='Оберіть файл КРАМ (.xlsx)',
                auto_upload=True,
                max_files=1,
                on_upload=handle_upload,
            ).props('accept=.xlsx').classes('flex-1')

            # Кнопки — праворуч
            with ui.row().classes('items-center gap-2 ml-auto flex-shrink-0'):
                process_btn = ui.button(
                    'Обробити',
                    icon='search',
                    on_click=lambda: ui.timer(0, _do_process, once=True),
                ).props('elevated')
                process_btn.disable()

                save_btn = ui.button(
                    'Зберегти зміни',
                    icon='cloud_upload',
                    on_click=lambda: ui.timer(0, _do_save, once=True),
                ).props('elevated color=orange')
                save_btn.set_visibility(False)

                export_btn = ui.button(
                    'Завантажити',
                    icon='download',
                    on_click=lambda: _export(state['rows']),
                ).props('elevated color=green')
                export_btn.set_visibility(False)

        # Статистичні картки
        stats_row = ui.row().classes('w-full gap-4')
        stats_row.set_visibility(False)

    # Контейнер таблиці — окремо від шапки, щоб скролилась незалежно
    results_container = ui.column().classes('w-full px-4 pb-4')

    # ------------------------------------------------------------------
    # Обробка файлу
    # ------------------------------------------------------------------
    async def _do_process():
        if not state['file_bytes']:
            ui.notify('Спочатку оберіть файл', type='warning')
            return

        process_btn.disable()
        save_btn.set_visibility(False)
        export_btn.set_visibility(False)
        results_container.clear()
        stats_row.set_visibility(False)
        state['pending'].clear()

        with results_container:
            ui.spinner(size='lg').classes('mt-6')
            ui.label('Обробляємо файл та звіряємо з базою...').classes('text-grey')

        try:
            rows: list[ErdrKramRow] = await auth_manager.execute(
                erdr_kram_ctrl.process_kram_file,
                auth_manager.get_current_context(),
                state['file_bytes'],
            )

            results_container.clear()

            if not rows:
                with results_container:
                    ui.label('У файлі не знайдено жодного рядка з даними.').classes('text-warning')
                return

            state['rows'] = rows

            total_count   = len(rows)
            found_count   = sum(1 for r in rows if r.found_in_db)
            erdr_count    = sum(1 for r in rows if r.found_in_db and r.db_erdr_date and r.db_erdr_date != NA)
            missing_count = sum(1 for r in rows if not r.found_in_db)
            pending_count = sum(1 for r in rows if '⚠️' in r.status)

            stats_row.clear()
            stats_row.set_visibility(True)
            with stats_row:
                _stat_card('Всього рядків',    total_count,   'blue-grey')
                _stat_card('Знайдено в базі',  found_count,   'positive')
                _stat_card('Є ЄРДР в базі',    erdr_count,    'positive')
                _stat_card('Не знайдено',       missing_count, 'negative')
                if pending_count:
                    _stat_card('Потребують запису', pending_count, 'warning')

            export_btn.set_visibility(True)

            with results_container:
                _render_table(rows)

        except Exception as e:
            results_container.clear()
            with results_container:
                ui.label(f'Помилка обробки: {e}').classes('text-negative')
            ui.notify(f'Помилка: {e}', type='negative')
        finally:
            process_btn.enable()

    # ------------------------------------------------------------------
    # Збереження ЄРДР у базу
    # ------------------------------------------------------------------
    async def _do_save():
        pending_rows = list(state['pending'].values())
        if not pending_rows:
            ui.notify('Немає змін для збереження', type='warning')
            return

        save_btn.disable()
        try:
            saved_count, errors = await auth_manager.execute(
                erdr_kram_ctrl.save_erdr_updates,
                auth_manager.get_current_context(),
                pending_rows,
            )

            if saved_count:
                ui.notify(f'Збережено {saved_count} ЄРДР-записів ✅', type='positive')
                state['pending'].clear()
                save_btn.set_visibility(False)

                # Оновлюємо статус збережених рядків у state
                saved_keys = {r.source_row for r in pending_rows}
                for r in state['rows']:
                    if r.source_row in saved_keys:
                        r.db_erdr_date     = r.erdr_date
                        r.db_erdr_notation = r.erdr_number
                        r.found_in_db      = True

                # Перемальовуємо таблицю
                results_container.clear()
                with results_container:
                    _render_table(state['rows'])

            if errors:
                for err in errors:
                    ui.notify(err, type='negative', timeout=8000)

        except Exception as e:
            ui.notify(f'Помилка збереження: {e}', type='negative')
        finally:
            save_btn.enable()

    # ------------------------------------------------------------------
    # Рендер таблиці
    # ------------------------------------------------------------------
    def _render_table(rows: list[ErdrKramRow]):
        columns = [
            {'name': 'row',        'label': '№',                 'field': 'row',        'align': 'center', 'sortable': True,  'style': 'width: 55px'},
            {'name': 'name',       'label': 'ПІБ (розпізнано)',  'field': 'name',       'align': 'left',   'sortable': True},
            {'name': 'birthday',   'label': 'Дата народження',   'field': 'birthday',   'align': 'center', 'sortable': True,  'style': 'width: 120px'},
            {'name': 'rnokpp',     'label': 'РНОКПП',            'field': 'rnokpp',     'align': 'center', 'sortable': True,  'style': 'width: 110px'},
            {'name': 'erdr_file',  'label': 'ЄРДР у файлі',      'field': 'erdr_file',  'align': 'left',   'sortable': True},
            {'name': 'status',     'label': 'Статус',             'field': 'status',     'align': 'left',   'sortable': True},
            {'name': 'db_erdr',    'label': 'ЄРДР в базі',       'field': 'db_erdr',    'align': 'left',   'sortable': True},
            {'name': 'db_erdr_dt', 'label': 'Дата ЄРДР (база)',  'field': 'db_erdr_dt', 'align': 'center', 'sortable': True,  'style': 'width: 120px'},
            {'name': 'actions',    'label': 'Дія',               'field': 'actions',    'align': 'center', 'style': 'width: 80px'},
            {'name': 'error',      'label': 'Помилка',            'field': 'error',      'align': 'left'},
        ]

        table_rows = []
        for r in rows:
            erdr_file_str = r.erdr_number if r.erdr_number != NA else '—'
            if r.erdr_date and r.erdr_date != NA:
                erdr_file_str += f' від {r.erdr_date}'

            in_pending = r.source_row in state['pending']

            table_rows.append({
                'id':         r.source_row,
                'row':        r.source_row,
                'name':       r.parsed_name     if r.parsed_name     != NA else '⚠️ не розпізнано',
                'birthday':   r.parsed_birthday if r.parsed_birthday != NA else '—',
                'rnokpp':     r.parsed_rnokpp   if r.parsed_rnokpp   != NA else '—',
                'erdr_file':  erdr_file_str,
                'status':     r.status,
                'db_erdr':    r.db_erdr_notation if r.db_erdr_notation != NA else '—',
                'db_erdr_dt': r.db_erdr_date     if r.db_erdr_date    != NA else '—',
                'error':      r.error or '',
                # службові поля для слота
                'has_warn':   '⚠️' in r.status,
                'is_queued':  in_pending,
            })

        ui.label(f'Результати: {len(table_rows)} записів').classes(
            'text-right text-grey w-full mb-1'
        )

        # Таблиця з повною шириною і власним вертикальним скролом
        table = ui.table(
            columns=columns,
            rows=table_rows,
            row_key='id',
        ).classes('w-full')
        table.props('bordered separator=cell flat dense virtual-scroll style="max-height: calc(100vh - 320px)"')

        # Слот для колонки «Статус» — кольоровий бейдж
        table.add_slot('body-cell-status', '''
            <q-td :props="props">
                <q-badge
                    :color="props.value.includes('✅') ? 'positive'
                          : props.value.includes('⚠️') ? 'warning'
                          : props.value.includes('❓') ? 'negative'
                          : 'grey'"
                    :label="props.value"
                    class="text-xs"
                />
            </q-td>
        ''')

        # Слот для колонки «Дія» — кнопка «Змінити» тільки для ⚠️ рядків
        table.add_slot('body-cell-actions', '''
            <q-td :props="props">
                <q-btn
                    v-if="props.row.has_warn && !props.row.is_queued"
                    flat dense round
                    icon="edit"
                    color="orange"
                    size="sm"
                    @click="() => $parent.$emit('queue_row', props.row)"
                >
                    <q-tooltip>Додати ЄРДР до збереження</q-tooltip>
                </q-btn>
                <q-icon
                    v-if="props.row.is_queued"
                    name="schedule"
                    color="orange"
                    size="sm"
                >
                    <q-tooltip>В черзі на збереження</q-tooltip>
                </q-icon>
            </q-td>
        ''')

        def _on_queue_row(e):
            row_id = e.args.get('id')
            # Знаходимо відповідний ErdrKramRow
            target = next((r for r in rows if r.source_row == row_id), None)
            if not target:
                return

            if row_id in state['pending']:
                # Повторне натискання — знімаємо з черги
                del state['pending'][row_id]
                ui.notify(f'Знято з черги: {target.parsed_name}', type='info')
            else:
                state['pending'][row_id] = target
                ui.notify(f'Додано до збереження: {target.parsed_name}', type='info')

            # Оновлюємо поле is_queued в рядку таблиці
            for tr in table.rows:
                if tr['id'] == row_id:
                    tr['is_queued'] = row_id in state['pending']
                    break
            table.update()

            save_btn.set_visibility(bool(state['pending']))

        table.on('queue_row', _on_queue_row)

    # ------------------------------------------------------------------
    # Статистична картка
    # ------------------------------------------------------------------
    def _stat_card(label: str, value: int, color: str):
        with ui.card().classes(f'p-3 text-center bg-{color}-50 flex-shrink-0'):
            ui.label(str(value)).classes('text-h5 font-bold')
            ui.label(label).classes('text-xs text-grey')

    # ------------------------------------------------------------------
    # Експорт у xlsx
    # ------------------------------------------------------------------
    def _export(rows: list[ErdrKramRow]):
        if not rows:
            ui.notify('Немає даних для експорту', type='warning')
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'ЄРДР КРАМ'

            border   = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'),  bottom=Side(style='thin'))
            c_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
            l_align  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
            bold     = Font(bold=True)
            hdr_fill = PatternFill('solid', fgColor='E0E0E0')
            fill_ok   = PatternFill('solid', fgColor='C6EFCE')
            fill_warn = PatternFill('solid', fgColor='FFEB9C')
            fill_err  = PatternFill('solid', fgColor='FFC7CE')

            headers = [
                '№ рядка', 'ПІБ (розпізнано)', 'Дата народження', 'РНОКПП',
                'ЄРДР у файлі', 'Дата ЄРДР (файл)',
                'Знайдено в базі', 'ЄРДР в базі (примітки)', 'Дата ЄРДР (база)',
                'Статус', 'Помилка парсингу',
            ]
            for col_num, hdr in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=hdr)
                cell.font = bold
                cell.fill = hdr_fill
                cell.alignment = c_align
                cell.border = border

            for row_idx, r in enumerate(rows, start=2):
                erdr_file_str = r.erdr_number if r.erdr_number != NA else ''
                if r.erdr_date and r.erdr_date != NA:
                    erdr_file_str += f' від {r.erdr_date}'

                vals = [
                    r.source_row,
                    r.parsed_name      if r.parsed_name      != NA else '',
                    r.parsed_birthday  if r.parsed_birthday  != NA else '',
                    r.parsed_rnokpp    if r.parsed_rnokpp    != NA else '',
                    erdr_file_str,
                    r.erdr_date        if r.erdr_date        != NA else '',
                    'Так' if r.found_in_db else 'Ні',
                    r.db_erdr_notation if r.db_erdr_notation != NA else '',
                    r.db_erdr_date     if r.db_erdr_date     != NA else '',
                    r.status,
                    r.error or '',
                ]

                if '✅' in r.status:
                    row_fill = fill_ok
                elif '⚠️' in r.status:
                    row_fill = fill_warn
                elif '❓' in r.status:
                    row_fill = fill_err
                else:
                    row_fill = None

                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = border
                    cell.alignment = c_align if col_idx not in (2, 5, 8, 10) else l_align
                    if row_fill:
                        cell.fill = row_fill

            ws.freeze_panes = 'A2'
            col_widths = [8, 32, 14, 13, 26, 14, 14, 26, 14, 28, 20]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            ui.download(buf.getvalue(), filename='ЄРДР_КРАМ_звірка.xlsx')
            ui.notify('Файл завантажується...', type='positive')

        except Exception as e:
            ui.notify(f'Помилка експорту: {e}', type='negative')