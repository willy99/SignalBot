from nicegui import ui, run
from dics.deserter_xls_dic import *
from gui.services.auth_manager import AuthManager

import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from domain.person_filter import PersonSearchFilter


def render_dbr_details_report_page(report_ctrl, person_ctrl, auth_manager: AuthManager):
    state = {'rows': [], 'columns': [], 'selected_year': None}

    year_options = person_ctrl.get_column_options().get(COLUMN_INSERT_DATE, [])

    async def on_year_change(e):
        if e.value:
            des_date_from.set_value(None)
            des_date_to.set_value(None)
            des_date_from.disable()
            des_date_to.disable()
        else:
            des_date_from.enable()
            des_date_to.enable()

    async def on_date_change(e):
        if des_date_from.value or des_date_to.value:
            des_year_filter.set_value(None)
            des_year_filter.disable()
        else:
            des_year_filter.enable()

    with ui.column().classes('w-full items-center p-4'):
        ui.label('Звіт: Список СЗЧ які очікують ЄРДР').classes('text-h4 mb-4')

        with ui.row().classes('w-full max-w-4xl items-center gap-4'):
            des_year_filter = ui.select(
                options=year_options,
                label='Рік СЗЧ',
                clearable=True
            ).classes('w-64').props('use-chips stack-label').on('update:model-value', lambda: do_report())

            des_date_from = ui.input('СЗЧ з (дата)', on_change=on_date_change).props('type=date clearable').classes(
                'w-40')
            des_date_to = ui.input('СЗЧ до (дата)', on_change=on_date_change).props('type=date clearable').classes(
                'w-40')

            search_btn = ui.button('Пошук', icon='search', on_click=lambda: do_report()).props('elevated')

            export_btn = ui.button('Експорт', icon='download', color='green',
                               on_click=lambda: export_to_excel(state['rows'], state['columns'],
                                                                state.get('selected_year')))

        results_container = ui.column().classes('w-full items-center mt-6')
        export_btn.bind_visibility_from(results_container, 'visible')


    async def do_report():
        des_year_val = des_year_filter.value if des_year_filter.value else None
        date_from_val = des_date_from.value if des_date_from.value else None
        date_to_val = des_date_to.value if des_date_to.value else None
        state['selected_year'] = des_year_val

        if not des_year_val and not date_from_val and not date_to_val:
            ui.notify('Введіть хоч якусь дату/рік для репорту', type='warning')
            return

        search_filter = PersonSearchFilter(
            des_year=[des_year_val] if des_year_val else [],
            des_date_from=date_from_val, des_date_to=date_to_val
        )

        des_year_filter.disable()
        des_date_from.disable()
        des_date_to.disable()
        search_btn.disable()
        export_btn.disable()

        with results_container:
            ui.spinner(size='lg').classes('mt-10')
            ui.label('Формування списку...').classes('text-grey')

        try:
            # Виклик методу вашого контролера. Змініть назву, якщо потрібно.
            data = await auth_manager.execute(report_ctrl.get_waiting_for_erdr_report, auth_manager.get_current_context(), search_filter)

            results_container.clear()

            if not data:
                ui.notify('За вказаними фільтрами даних не знайдено.', type='warning')
                return

            with results_container:
                rows, columns = results_ui(data)
                state['rows'] = rows
                state['columns'] = columns

        except Exception as e:
            results_container.clear()
            ui.notify(f'Помилка формування звіту: {e}', type='negative')
        finally:
            des_year_filter.enable()
            des_date_from.enable()
            des_date_to.enable()
            search_btn.enable()
            export_btn.enable()


def results_ui(data):
    if not data:
        return [], []

    # Визначаємо колонки для таблиці та Excel
    columns = [
        {'name': 'des_date', 'label': COLUMN_INSERT_DATE, 'field': 'des_date', 'align': 'center', 'sortable': True, 'headerClasses': 'bg-blue-100'},
        {'name': 'pib', 'label': COLUMN_NAME, 'field': 'pib', 'align': 'left', 'sortable': True, 'headerClasses': 'bg-blue-100'},
        {'name': 'rnokpp', 'label': COLUMN_ID_NUMBER, 'field': 'rnokpp', 'align': 'center', 'sortable': True, 'headerClasses': 'bg-blue-100'},
        {'name': 'dob', 'label': COLUMN_BIRTHDAY, 'field': 'dob', 'align': 'center', 'sortable': True, 'headerClasses': 'bg-blue-100'},
        {'name': 'rank', 'label': COLUMN_TITLE, 'field': 'rank', 'align': 'left', 'sortable': True, 'headerClasses': 'bg-blue-100'},
        {'name': 'unit', 'label': COLUMN_SUBUNIT, 'field': 'unit', 'align': 'left', 'sortable': True, 'headerClasses': 'bg-blue-100'},
        {'name': 'dbr_date', 'label': COLUMN_DBR_DATE, 'field': 'dbr_date', 'align': 'center',
         'sortable': True, 'headerClasses': 'bg-blue-100'},
        {'name': 'dbr_number', 'label': COLUMN_DBR_NUMBER, 'field': 'dbr_number', 'align': 'center',
         'sortable': True, 'headerClasses': 'bg-blue-100'},
        {'name': 'erdr_date', 'label': COLUMN_ERDR_DATE, 'field': 'erdr_date', 'align': 'center',
         'sortable': True, 'headerClasses': 'bg-blue-100'},
        {'name': 'erdr_number', 'label': COLUMN_ERDR_NOTATION, 'field': 'erdr_number', 'align': 'center',
         'sortable': True, 'headerClasses': 'bg-blue-100'},
    ]

    # Якщо контролер повертає об'єкти (напр. Person), конвертуємо в словники.
    # Якщо вже повертає словники, просто адаптуємо ключі.
    rows = []
    for i, item in enumerate(data):
        # Приклад мапінгу (адаптуйте під ключі, які реально повертає ваш контролер)
        row = {
            'id': i,
            'des_date': item.get('des_date', ''),
            'pib': item.get('pib', ''),
            'rnokpp': item.get('rnokpp', ''),
            'dob': item.get('dob', ''),
            'rank': item.get('rank', ''),
            'unit': item.get('unit', ''),
            'dbr_date': item.get('dbr_date', ''),
            'dbr_number': item.get('dbr_number', ''),
            'erdr_date': item.get('erdr_date', ''),
            'erdr_number': item.get('erdr_number', ''),
        }
        rows.append(row)

    ui.label(f'Знайдено записів: {len(rows)}').classes('w-full text-right text-gray-500 font-bold mb-2')

    table = ui.table(columns=columns, rows=rows, row_key='id').classes('w-full max-w-9xl')
    table.props('bordered separator=cell flat dense')

    return rows, columns


def export_to_excel(rows, columns, year_suffix=None):
    if not rows:
        ui.notify('Немає даних для експорту', type='warning')
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Список СЗЧ - чекаємо ЄРДР"

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        bold_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="E0E0E0")

        # 1. ЗАПИСУЄМО ЗАГОЛОВКИ
        for col_num, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_num, value=col['label'])
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border

        # 2. ЗАПИСУЄМО РЯДКИ
        for row_idx, r in enumerate(rows, start=2):
            for col_idx, col in enumerate(columns, start=1):
                val = r.get(col['field'], '')
                if val is None:
                    val = ''

                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

                # ПІБ, Звання та Підрозділ вирівнюємо по лівому краю, інше - по центру
                if col['field'] in ['pib', 'unit', 'rank']:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

        # 3. ЗАКРІПЛЮЄМО ШАПКУ
        ws.freeze_panes = 'A2'

        # 4. АВТОПІДБІР ШИРИНИ СТОВПЦІВ
        widths = {
            'des_date': 10,
            'pib': 30,
            'rnokpp': 15,
            'dob': 10,
            'rank': 15,
            'unit': 25,
            'dbr_date': 12,
            'dbr_number': 15,
            'erdr_date': 12,
            'erdr_number': 15
        }

        for col_idx, col in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = widths.get(col['field'], 15)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        if year_suffix:
            final_filename = f'Список СЗЧ - чекаємо ЄРДР ({year_suffix}).xlsx'
        else:
            final_filename = 'Список СЗЧ - чекаємо ЄРДР.xlsx'

        ui.download(buffer.getvalue(), filename=final_filename)
        ui.notify('Файл завантажується...', type='positive')

    except Exception as e:
        print(f"Excel Error Trace: {e}")
        ui.notify(f'Помилка генерації файлу: {e}', type='negative')