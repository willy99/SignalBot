from nicegui import ui, run
from gui.services.request_context import RequestContext
from domain.person_filter import PersonSearchFilter
from dics.deserter_xls_dic import REVIEW_STATUS_NOT_ASSIGNED, REVIEW_STATUS_EXECUTING, REVIEW_STATUS_CLOSED

import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


def render_yearly_report_page(report_ctrl, ctx: RequestContext):
    state = {'rows': [], 'columns': []}

    with ui.column().classes('w-full items-center p-4'):
        ui.label('Звіт: Статистика СЗЧ по роках').classes('text-h4 mb-4')

        with ui.row().classes('w-full max-w-4xl justify-center items-center gap-4 mb-6'):
            generate_btn = ui.button('Сформувати звіт', icon='analytics', on_click=lambda: do_report()) \
                .props('elevated size="lg" color="primary"')

            export_btn = ui.button('Експорт в Excel', icon='download', color='green',
                                   on_click=lambda: export_to_excel(state['rows'], state['columns'])) \
                .props('elevated size="lg"')
            export_btn.set_visibility(False)

        results_container = ui.column().classes('w-full items-center mt-2')

    async def do_report():
        generate_btn.disable()
        export_btn.set_visibility(False)

        with results_container:
            ui.spinner(size='xl').classes('mt-10')
            ui.label('Аналіз бази даних... Це може зайняти трохи часу.').classes('text-grey text-lg mt-2')

        try:
            # Передаємо порожній фільтр, бо нам потрібні всі роки
            empty_filter = PersonSearchFilter()
            data = await run.io_bound(report_ctrl.get_yearly_desertion_report, ctx, empty_filter)

            results_container.clear()

            if not data:
                ui.notify('Немає даних для побудови звіту.', type='warning')
                return

            # Малюємо таблицю
            with results_container:
                rows, columns = results_ui(data)
                state['rows'] = rows
                state['columns'] = columns
                export_btn.set_visibility(True)

        except Exception as e:
            results_container.clear()
            import traceback
            traceback.print_exc()
            ui.notify(f'Помилка формування звіту: {e}', type='negative')
        finally:
            generate_btn.enable()


def results_ui(data):
    if not data:
        return [], []

    rows = []

    # 1. Створюємо об'єкт для глобального підсумку (Разом за всі роки)
    grand_total = {
        'year': 'ВСЬОГО ЗА ВСІ РОКИ',
        's_under': 0, 'o_under': 0, 's_over': 0, 'o_over': 0,
        's_ret_mu': 0, 's_ret_res': 0, 'o_ret': 0,
        's_diff': 0, 'o_diff': 0,
        'expl_not_ass': 0, 'expl_not_exe': 0, 'expl_not_cls': 0,
        's_dupl': 0, 'o_dupl': 0,
        'un_sold_des': 0, 'un_sold_ret': 0,
        'un_serg_des': 0, 'un_serg_ret': 0,
        'un_offc_des': 0, 'un_offc_ret': 0,
        'st_term': 0, 'st_call': 0, 'st_contr': 0,
        'pl_ppd': 0, 'pl_rvbz': 0, 'pl_other': 0,
        'weapon': 0,
        'rev_specified_total': 0, 'rev_specified_of': 0,
        'rev_dbr_notif_total': 0, 'rev_dbr_notif_of': 0,
        'rev_dbr_mater_total': 0, 'rev_dbr_mater_of': 0,
        'rev_dbr_nonerdr_total': 0, 'rev_dbr_nonerdr_of': 0,
        'rev_dbr_erdr_total': 0, 'rev_dbr_erdr_of': 0,
        'rev_suspend_total': 0, 'rev_suspend_of': 0,
        'rev_courts_total': 0, 'rev_courts_of': 0,
        'rev_punish_total': 0, 'rev_punish_of': 0,
        'rev_nonevil_total': 0, 'rev_nonevil_of': 0,
        'is_grand_total': True
    }

    # Сортуємо роки. "Невідомо" відправляємо в кінець списку.
    sorted_years = list(data.keys())
    sorted_years.sort(key=lambda y: (y == "Невідомо", y))

    for year in sorted_years:
        stats = data[year]

        # Підрахунок полів для конкретного року
        s_under = stats['рядовий_сержант']['under_3'] + stats['офіцер']['under_3']
        s_over = stats['рядовий_сержант']['over_3'] + stats['офіцер']['over_3']
        o_under = stats['офіцер']['under_3']
        o_over = stats['офіцер']['over_3']
        s_ret_mu = stats['рядовий_сержант']['ret_mu']
        s_ret_res = stats['рядовий_сержант']['ret_res']
        o_ret_mu = stats['офіцер']['ret_mu']
        s_dupl = stats['рядовий']['dupl'] + stats['сержант']['dupl'] + stats['офіцер']['dupl']
        o_dupl = stats['офіцер']['dupl']

        un_sold_des = stats['рядовий']['un_des']
        un_sold_ret = stats['рядовий']['un_ret']
        un_serg_des = stats['сержант']['un_des']
        un_serg_ret = stats['сержант']['un_ret']
        un_offc_des = stats['офіцер']['un_des']
        un_offc_ret = stats['офіцер']['un_ret']

        st_term = stats['all']['st_term']
        st_call = stats['all']['st_call']
        st_contr = stats['all']['st_contr']
        pl_ppd = stats['all']['pl_ppd']
        pl_rvbz = stats['all']['pl_rvbz']
        pl_other = stats['all']['pl_other']

        expl_not_ass = stats['all'].get(REVIEW_STATUS_NOT_ASSIGNED, 0)
        expl_not_exe = stats['all'].get(REVIEW_STATUS_EXECUTING, 0)
        expl_not_cls = stats['all'].get(REVIEW_STATUS_CLOSED, 0)
        weapon = stats['all'].get('weapon', 0)

        def get_rev_stat(key):
            s_val = stats['all'].get(key, 0)
            o_val = stats['офіцер'].get(key, 0)
            return (s_val + o_val), o_val

        rev_specified_total, rev_specified_of = get_rev_stat('rev_specified')
        rev_dbr_notif_total, rev_dbr_notif_of = get_rev_stat('rev_dbr_notif')
        rev_dbr_mater_total, rev_dbr_mater_of = get_rev_stat('rev_dbr_mater')
        rev_dbr_nonerdr_total, rev_dbr_nonerdr_of = get_rev_stat('rev_dbr_nonerdr')
        rev_dbr_erdr_total, rev_dbr_erdr_of = get_rev_stat('rev_dbr_erdr')
        rev_suspend_total, rev_suspend_of = get_rev_stat('rev_suspend')
        rev_courts_total, rev_courts_of = get_rev_stat('rev_courts')
        rev_punish_total, rev_punish_of = get_rev_stat('rev_punish')
        rev_nonevil_total, rev_nonevil_of = get_rev_stat('rev_nonevil')

        # Динамічні поля для цього року
        s_diff = s_under + s_over - s_ret_mu - s_ret_res
        o_diff = o_under + o_over - o_ret_mu

        row = {
            'year': year,
            's_under': s_under, 'o_under': o_under, 's_over': s_over, 'o_over': o_over,
            's_total': s_under + s_over, 'o_total': o_under + o_over,
            's_ret_mu': s_ret_mu, 's_ret_res': s_ret_res,
            's_total_ret': s_ret_mu + s_ret_res, 'o_ret': o_ret_mu,
            's_diff': s_diff, 'o_diff': o_diff,
            'expl_not_ass': expl_not_ass, 'expl_not_exe': expl_not_exe, 'expl_not_cls': expl_not_cls,
            's_dupl': s_dupl, 'o_dupl': o_dupl,
            'un_sold_des': un_sold_des, 'un_sold_ret': un_sold_ret,
            'un_serg_des': un_serg_des, 'un_serg_ret': un_serg_ret,
            'un_offc_des': un_offc_des, 'un_offc_ret': un_offc_ret,
            'st_term': st_term, 'st_call': st_call, 'st_contr': st_contr,
            'pl_ppd': pl_ppd, 'pl_rvbz': pl_rvbz, 'pl_other': pl_other,
            'weapon': weapon,
            'rev_specified_total': rev_specified_total, 'rev_specified_of': rev_specified_of,
            'rev_dbr_notif_total': rev_dbr_notif_total, 'rev_dbr_notif_of': rev_dbr_notif_of,
            'rev_dbr_mater_total': rev_dbr_mater_total, 'rev_dbr_mater_of': rev_dbr_mater_of,
            'rev_dbr_nonerdr_total': rev_dbr_nonerdr_total, 'rev_dbr_nonerdr_of': rev_dbr_nonerdr_of,
            'rev_dbr_erdr_total': rev_dbr_erdr_total, 'rev_dbr_erdr_of': rev_dbr_erdr_of,
            'rev_suspend_total': rev_suspend_total, 'rev_suspend_of': rev_suspend_of,
            'rev_courts_total': rev_courts_total, 'rev_courts_of': rev_courts_of,
            'rev_punish_total': rev_punish_total, 'rev_punish_of': rev_punish_of,
            'rev_nonevil_total': rev_nonevil_total, 'rev_nonevil_of': rev_nonevil_of,
            'is_grand_total': False
        }
        rows.append(row)

        # ДОДАЄМО ДО ГЛОБАЛЬНОГО ПІДСУМКУ
        for key in grand_total.keys():
            if key not in ['year', 'is_grand_total', 's_total', 'o_total', 's_total_ret']:
                grand_total[key] += row[key]

    # Фінальні розрахунки для глобального підсумку
    grand_total['s_total'] = grand_total['s_under'] + grand_total['s_over']
    grand_total['o_total'] = grand_total['o_under'] + grand_total['o_over']
    grand_total['s_total_ret'] = grand_total['s_ret_mu'] + grand_total['s_ret_res']
    grand_total['s_diff'] = grand_total['s_total'] - grand_total['s_total_ret']
    grand_total['o_diff'] = grand_total['o_total'] - grand_total['o_ret']

    # Вставляємо підсумок на самий початок
    rows.insert(0, grand_total)

    # Визначення колонок
    columns = [
        {'name': 'year', 'label': 'Рік', 'field': 'year', 'align': 'left'},
        {'name': 's_under', 'label': '< 3 (Всього)', 'field': 's_under'},
        {'name': 'o_under', 'label': '< 3 (Офіц.)', 'field': 'o_under'},
        {'name': 's_over', 'label': '>= 3 (Всього)', 'field': 's_over'},
        {'name': 'o_over', 'label': '>= 3 (Офіц.)', 'field': 'o_over'},
        {'name': 's_total', 'label': 'Всього', 'field': 's_total', 'headerClasses': 'bg-blue-100'},
        {'name': 'o_total', 'label': 'Всього (Офіц.)', 'field': 'o_total', 'headerClasses': 'bg-blue-100'},
        {'name': 's_ret_mu', 'label': 'У В/Ч (С/С)', 'field': 's_ret_mu', 'headerClasses': 'bg-unit-return'},
        {'name': 's_ret_res', 'label': 'В БРез (С/С)', 'field': 's_ret_res', 'headerClasses': 'bg-unit-return'},
        {'name': 's_total_ret', 'label': 'Всього (С/С)', 'field': 's_total_ret', 'headerClasses': 'bg-blue-100'},
        {'name': 'o_ret', 'label': 'В т.ч. Офіц.', 'field': 'o_ret', 'headerClasses': 'bg-unit-return'},
        {'name': 's_diff', 'label': 'Кількість СЗЧ (С/С)', 'field': 's_diff', 'headerClasses': 'bg-blue-100'},
        {'name': 'o_diff', 'label': 'Кількість СЗЧ (Офіц.)', 'field': 'o_diff', 'headerClasses': 'bg-blue-100'},
        {'name': 'expl_not_ass', 'label': 'Не призначено', 'field': 'expl_not_ass'},
        {'name': 'expl_not_exe', 'label': 'Проводяться', 'field': 'expl_not_exe'},
        {'name': 'expl_not_cls', 'label': 'Завершено', 'field': 'expl_not_cls'},
        {'name': 's_dupl', 'label': 'Всього', 'field': 's_dupl'},
        {'name': 'o_dupl', 'label': 'Офіцерів', 'field': 'o_dupl'},
        {'name': 'un_sold_des', 'label': 'СЗЧ Солдат', 'field': 'un_sold_des'},
        {'name': 'un_sold_ret', 'label': 'Пов. Солдат', 'field': 'un_sold_ret', 'headerClasses': 'bg-unit-return'},
        {'name': 'un_serg_des', 'label': 'СЗЧ Сержант', 'field': 'un_serg_des'},
        {'name': 'un_serg_ret', 'label': 'Пов. Сержант', 'field': 'un_serg_ret', 'headerClasses': 'bg-unit-return'},
        {'name': 'un_offc_des', 'label': 'СЗЧ Офіцер', 'field': 'un_offc_des'},
        {'name': 'un_offc_ret', 'label': 'Пов. Офіцер', 'field': 'un_offc_ret', 'headerClasses': 'bg-unit-return'},
        {'name': 'st_term', 'label': 'Строкова служба', 'field': 'st_term'},
        {'name': 'st_call', 'label': 'За призивом', 'field': 'st_call'},
        {'name': 'st_contr', 'label': 'За контрактом', 'field': 'st_contr'},
        {'name': 'pl_ppd', 'label': 'З ППД', 'field': 'pl_ppd'},
        {'name': 'pl_rvbz', 'label': 'З РВБЗ', 'field': 'pl_rvbz'},
        {'name': 'pl_other', 'label': 'Інше', 'field': 'pl_other'},
        {'name': 'weapon', 'label': 'Зброя', 'field': 'weapon'},
        {'name': 'rev_specified_total', 'label': 'Всього', 'field': 'rev_specified_total'},
        {'name': 'rev_specified_of', 'label': 'В т.ч офіцерів', 'field': 'rev_specified_of'},
        {'name': 'rev_dbr_notif_total', 'label': 'Всього', 'field': 'rev_dbr_notif_total'},
        {'name': 'rev_dbr_notif_of', 'label': 'В т.ч офіцерів', 'field': 'rev_dbr_notif_of'},
        {'name': 'rev_dbr_mater_total', 'label': 'Всього', 'field': 'rev_dbr_mater_total'},
        {'name': 'rev_dbr_mater_of', 'label': 'В т.ч офіцерів', 'field': 'rev_dbr_mater_of'},
        {'name': 'rev_dbr_nonerdr_total', 'label': 'Всього', 'field': 'rev_dbr_nonerdr_total'},
        {'name': 'rev_dbr_nonerdr_of', 'label': 'В т.ч офіцерів', 'field': 'rev_dbr_nonerdr_of'},
        {'name': 'rev_dbr_erdr_total', 'label': 'Всього', 'field': 'rev_dbr_erdr_total'},
        {'name': 'rev_dbr_erdr_of', 'label': 'В т.ч офіцерів', 'field': 'rev_dbr_erdr_of'},
        {'name': 'rev_suspend_total', 'label': 'Всього', 'field': 'rev_suspend_total'},
        {'name': 'rev_suspend_of', 'label': 'В т.ч офіцерів', 'field': 'rev_suspend_of'},
        {'name': 'rev_courts_total', 'label': 'Всього', 'field': 'rev_courts_total'},
        {'name': 'rev_courts_of', 'label': 'В т.ч офіцерів', 'field': 'rev_courts_of'},
        {'name': 'rev_punish_total', 'label': 'Всього', 'field': 'rev_punish_total'},
        {'name': 'rev_punish_of', 'label': 'В т.ч офіцерів', 'field': 'rev_punish_of'},
        {'name': 'rev_nonevil_total', 'label': 'Всього', 'field': 'rev_nonevil_total'},
        {'name': 'rev_nonevil_of', 'label': 'В т.ч офіцерів', 'field': 'rev_nonevil_of'},
    ]

    table = ui.table(columns=columns, rows=rows, row_key='year').classes('w-full report-table')
    table.props('bordered separator=cell flat')

    # Кастомний слот шапки (відрізняється від попереднього лише першим colspan="1")
    table.add_slot('header', '''
            <q-tr>
                <q-th colspan="1" class="text-center bg-grey-3 text-bold text-subtitle1">Рік</q-th>
                <q-th colspan="6" class="text-center bg-grey-3 text-bold text-subtitle1">Випадки СЗЧ</q-th>
                <q-th colspan="4" class="text-center bg-grey-3 text-bold text-subtitle1">Повернення</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Особового складу в СЗЧ</q-th>
                <q-th colspan="3" class="text-center bg-grey-3 text-bold text-subtitle1">Відпрацювання службових розслідувань</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">СЗЧ 2 і більше</q-th>
                <q-th colspan="6" class="text-center bg-grey-3 text-bold text-subtitle1">Унікальні випадки СЗЧ та повернення</q-th>
                <q-th colspan="3" class="text-center bg-grey-3 text-bold text-subtitle1">Вид служби</q-th>
                <q-th colspan="3" class="text-center bg-grey-3 text-bold text-subtitle1">Місце скоєння</q-th>
                <q-th colspan="1" class="text-center bg-grey-3 text-bold text-subtitle1">Зі зброєю</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Виведено у розпорядження</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Повідомлень про КП до ВСП, ДБР</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Матеріалів с/р до ВСП, ДБР</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Не отримано витяг з ЄРДР</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Отримано витяг з ЄРДР</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Призупинена в/служба</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Кількість вироків судів</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Відбувають покарання</q-th>
                <q-th colspan="2" class="text-center bg-grey-3 text-bold text-subtitle1">Не є суб'єктом злочину</q-th>
            </q-tr>
            <q-tr>
                <q-th v-for="col in props.cols" :key="col.name" :props="props" :class="col.header_classes">
                    {{ col.label }}
                </q-th>
            </q-tr>
        ''')

    # Кастомне відображення рядків (з підсвіткою "Разом")
    table.add_slot('body', '''
            <q-tr :props="props" :class="props.row.is_grand_total ? 'bg-grand-total font-bold' : ''">
                <q-td v-for="col in props.cols" :key="col.name" :props="props" 
                    :class="[
                        (col.name === 's_total' || col.name === 'o_total' || col.name === 's_total_ret' || col.name === 's_diff' || col.name === 'o_diff') ? 'bg-unit-total' : '',
                        (['s_ret_mu', 's_ret_res', 'o_ret', 'un_sold_ret', 'un_serg_ret', 'un_offc_ret'].includes(col.name)) ? 'bg-unit-return' : ''
                    ]">
                    <template v-if="col.name === 'year' && props.row.is_grand_total">
                        <q-icon name="star" color="orange" /> {{ col.value }}
                    </template>
                    <template v-else-if="col.name === 'year'">
                        <b>{{ col.value }}</b>
                    </template>
                    <template v-else>
                        {{ col.value }}
                    </template>
                </q-td>
            </q-tr>
        ''')

    return rows, columns


def export_to_excel(rows, columns):
    if not rows:
        ui.notify('Немає даних для експорту', type='warning')
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Звіт по роках"

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bold_font = Font(bold=True)

        header_fill = PatternFill("solid", fgColor="EEEEEE")
        header_row2_fill = PatternFill("solid", fgColor="E0E0E0")
        grand_total_fill = PatternFill("solid", fgColor="FFE0B2")
        blue_col_fill = PatternFill("solid", fgColor="BBDEFB")
        green_col_fill = PatternFill("solid", fgColor="C1F5DD")

        # 1. СТВОРЕННЯ ШАПКИ (Зі зміщенням вліво порівняно зі звітом по підрозділах)
        def set_header(range_str, text):
            ws.merge_cells(range_str)
            top_left_cell = range_str.split(':')[0]
            ws[top_left_cell].value = text
            for row in ws[range_str]:
                for cell in row:
                    cell.fill = header_fill
                    cell.font = bold_font
                    cell.alignment = center_align
                    cell.border = thin_border

        set_header('A1:A1', 'Рік')
        set_header('B1:G1', 'Випадки СЗЧ')
        set_header('H1:K1', 'Повернення')
        set_header('L1:M1', 'Особового складу в СЗЧ')
        set_header('N1:P1', 'Відпрацювання службових розслідувань')
        set_header('Q1:R1', 'СЗЧ 2 і більше')
        set_header('S1:X1', 'Унікальні випадки СЗЧ та повернення')
        set_header('Y1:AA1', 'Вид служби')
        set_header('AB1:AD1', 'Місце скоєння')
        set_header('AE1:AE1', 'Зі зброєю')
        set_header('AF1:AG1', 'Виведено у розпорядження')
        set_header('AH1:AI1', 'Повідомлень про КП до ВСП, ДБР')
        set_header('AJ1:AK1', 'Матеріалів с/р до ВСП, ДБР')
        set_header('AL1:AM1', 'Не отримано витяг з ЄРДР')
        set_header('AN1:AO1', 'Отримано витяг з ЄРДР')
        set_header('AP1:AQ1', 'Призупинена в/служба')
        set_header('AR1:AS1', 'Кількість вироків судів')
        set_header('AT1:AU1', 'Відбувають покарання')
        set_header('AV1:AW1', "Не є суб'єктом злочину")

        # 2. ДРУГИЙ РЯДОК ЗАГОЛОВКІВ
        blue_cols = ['s_total', 'o_total', 's_total_ret', 's_diff', 'o_diff']

        for col_num, col in enumerate(columns, start=1):
            cell = ws.cell(row=2, column=col_num, value=col['label'])
            cell.fill = header_row2_fill
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border

        # 3. ЗАПИС ДАНИХ
        for row_idx, r in enumerate(rows, start=3):
            is_grand_total = r.get('is_grand_total')

            for col_idx, col in enumerate(columns, start=1):
                val = r.get(col['field'])
                if val is None:
                    val = ''
                elif not isinstance(val, (int, float)):
                    val = str(val)

                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

                if is_grand_total:
                    cell.fill = grand_total_fill
                    cell.font = bold_font
                else:
                    if col['name'] in blue_cols:
                        cell.fill = blue_col_fill
                    elif 'bg-unit-return' in str(col.get('headerClasses', '')):
                        cell.fill = green_col_fill

        # 4. ЗАКРІПЛЕННЯ ОБЛАСТЕЙ
        # 'B3' означає закріплення 1 першої колонки і 2 верхніх рядків
        ws.freeze_panes = 'B3'

        # 5. ШИРИНА КОЛОНОК
        ws.column_dimensions['A'].width = 18
        for col_idx in range(2, len(columns) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 12

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        ui.download(buffer.getvalue(), filename='Звіт_по_роках.xlsx')
        ui.notify('Файл завантажується...', type='positive')

    except Exception as e:
        import traceback
        traceback.print_exc()
        ui.notify(f'Помилка генерації файлу: {e}', type='negative')