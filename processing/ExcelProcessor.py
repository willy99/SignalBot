import openpyxl
import io
import warnings
from copy import copy
from config import DESERTER_TAB_NAME
from dics.deserter_xls_dic import *
from dics.deserter_xls_dic import NA
from typing import List, Dict, Any
from storage.StorageFactory import StorageFactory
from utils.utils import format_ukr_date, get_typed_value

class ExcelProcessor:
    def __init__(self, file_path, batch_processing=False):
        self.file_path: str = file_path
        self.workbook = None
        self.sheet = None
        self.column_map: Dict[str, int] = {}  # {назва: номер_колонки}
        self.batch_processing = batch_processing
        self.fileProxy = StorageFactory.create_client(file_path)

        warnings.filterwarnings("ignore", category=UserWarning)
        with self.fileProxy as smb:
            self._load_workbook(smb) # reading file to check it exists
            if not self.batch_processing:
                self.workbook = None
            self.file_buffer = None

    def upsert_record(self, records_list: List[Dict[str, Any]]) -> None:
        if not records_list:
            return
        # Якщо ми не в батчі і воркбук не завантажений - завантажуємо
        if self.workbook is None:
            with self.fileProxy as smb:
                self._load_workbook(smb)

        self._processRow(records_list)
        # Якщо НЕ батч - зберігаємо негайно
        if not self.batch_processing:
            with self.fileProxy as smb:
                self.save(smb)

        # print(f"Додано запис у рядок {next_row}")

    def _processRow(self, records_list):
        id_col_idx = self.column_map.get(COLUMN_INCREMEMTAL.lower())
        if not id_col_idx:
            print("❌ Помилка: Не знайдено колонку №")
            return
        target_insert_row = 2
        for row in range(2, self.sheet.max_row + 2):
            cell_val = self.sheet.cell(row=row, column=id_col_idx).value
            if cell_val is None or str(cell_val).strip() == "":
                target_insert_row = row
                break

        # 3. Визначаємо останній існуючий ID (з рядка над target_insert_row)
        last_val = self.sheet.cell(row=target_insert_row - 1, column=id_col_idx).value
        max_col = len(self.column_map) if self.column_map else self.sheet.max_column
        try:
            current_id = int(last_val) if last_val and str(last_val).isdigit() else 0
        except (ValueError, TypeError):
            current_id = 0

        # 2. Перебір кожного словника в масиві
        for data_dict in records_list:
            # пошук чувака в базі

            existing_row = self._find_existing_row(data_dict)
            if existing_row:
                # Дивимося на дати повернення та сзч
                # Логіка оновлення:: якщо дата повернення порожня, а це довідка повернення - пхаємо цю дату
                for col_name, value in data_dict.items():
                    idx = self.column_map.get(col_name.lower())
                    if idx:
                        cell = self.sheet.cell(row=existing_row, column=idx)
                        # Оновлюємо тільки якщо в базі пусто, а в нових даних щось є
                        if (not cell.value or cell.value == NA) and value:
                            cell.value = get_typed_value(value)
                            print('--- оновлюємо ' + str(value))
            else:
                current_id += 1
                # Вставляємо новий порожній рядок
                self.sheet.insert_rows(target_insert_row)
                sample_row = target_insert_row - 1 if target_insert_row > 2 else 2

                for col_idx in range(1, max_col + 1):
                    new_cell = self.sheet.cell(row=target_insert_row, column=col_idx)
                    old_cell = self.sheet.cell(row=sample_row, column=col_idx)

                    # Копіюємо стилі
                    if old_cell.has_style:
                        new_cell.font = copy(old_cell.font)
                        new_cell.border = copy(old_cell.border)
                        new_cell.number_format = copy(old_cell.number_format)

                        new_alignment = copy(old_cell.alignment)
                        new_alignment.wrapText = False  # Один рядок
                        new_alignment.vertical = 'center'
                        new_cell.alignment = new_alignment

                # 3. Записуємо ID та дані
                self.sheet.cell(row=target_insert_row, column=1).value = current_id

                for col_name, value in data_dict.items():
                    idx = self.column_map.get(col_name.lower())
                    if idx:
                        self.sheet.cell(row=target_insert_row, column=idx).value = get_typed_value(value)

                # 4. Фіксуємо висоту
                self.sheet.row_dimensions[target_insert_row].height = 15

                # Переходимо до наступного рядка для наступного словника
                target_insert_row += 1

    def _find_existing_row(self, data_dict: Dict[str, Any]):
        """Шукає номер рядка за ПІБ, Датою народження та РНОКПП."""
        pib = str(data_dict.get(COLUMN_NAME, '')).strip().lower()
        dob = str(data_dict.get(COLUMN_BIRTHDAY, '')).strip()
        rnokpp = str(data_dict.get(COLUMN_ID_NUMBER, '')).strip()
        des_date = str(data_dict.get(COLUMN_DESERTION_DATE, '')).strip()
        ret_date = str(data_dict.get(COLUMN_RETURN_DATE, '')).strip()
        ret_reserve_date = str(data_dict.get(COLUMN_RETURN_TO_RESERVE_DATE, '')).strip()

        pid_col = self.column_map.get(COLUMN_INCREMEMTAL.lower())
        pib_col = self.column_map.get(COLUMN_NAME.lower())
        dob_col = self.column_map.get(COLUMN_BIRTHDAY.lower())
        rnokpp_col = self.column_map.get(COLUMN_ID_NUMBER.lower())
        des_date_col = self.column_map.get(COLUMN_DESERTION_DATE.lower())
        ret_date_col = self.column_map.get(COLUMN_RETURN_DATE.lower())
        ret_reserve_date_col = self.column_map.get(COLUMN_RETURN_TO_RESERVE_DATE.lower())

        print('--- 🔎: Пошук чувака в базі:: ' + str(pib) + ' || ' + str(dob) + ' || ' + str(rnokpp) + '; сзч||взад:' + str(des_date) + ' || ' + str(ret_date))
        if not all([pib_col, dob_col, rnokpp_col, des_date_col, ret_date_col]):
            return None
        last_found = None

        for row in range(2, self.sheet.max_row + 1):
            s_pid = str(self.sheet.cell(row=row, column=pid_col).value or "").strip().lower()
            s_pib = str(self.sheet.cell(row=row, column=pib_col).value or "").strip().lower()
            s_dob = format_ukr_date(str(self.sheet.cell(row=row, column=dob_col).value or "").strip())
            s_rnokpp = str(self.sheet.cell(row=row, column=rnokpp_col).value or "").strip()
            s_des_date = format_ukr_date(str(self.sheet.cell(row=row, column=des_date_col).value or "").strip())
            s_ret_date = format_ukr_date(str(self.sheet.cell(row=row, column=ret_date_col).value or "").strip())
            s_ret_reserve_date = format_ukr_date(str(self.sheet.cell(row=row, column=ret_reserve_date_col).value or "").strip())
            # todo if 12/31/20 - КОСТИЛЬ!
            if s_ret_date == '31.12.2020':
                s_ret_date = ''
                self.sheet.cell(row=row, column=ret_date_col).value = None
            if s_ret_reserve_date == '31.12.2020':
                s_ret_reserve_date = ''
                self.sheet.cell(row=row, column=ret_reserve_date_col).value = None

            if s_pib == pib and s_dob == dob and s_rnokpp == rnokpp:
                print('--- ID: ' + str(s_pid) + ' des_date='+str(s_des_date))
                if des_date == s_des_date or (s_ret_date == "" and s_ret_reserve_date == ""):
                    print('--- 🔎⚠️: Чувак вже в базі, будемо доповнювати запис! (ID:' + s_pid + ')')
                    return row
                # last_found = row
        print('--- 🔎➕: Чувака немає, додаємо')
        return last_found

    def _find_last_row(self):
        return self.sheet.max_row

    def _build_column_map(self):
        """Створює словник імен колонок для швидкого доступу"""
        if self.sheet:
            header_row = next(self.sheet.iter_rows(min_row=1, max_row=1))
            for cell in header_row:
                if cell.value:
                    clean_name = str(cell.value).strip().lower()
                    self.column_map[clean_name] = cell.column


    def _load_workbook(self, fileProxy) -> None:
        try:
            print(f'>> LOADING WORKBOOK...')
            self.file_buffer = fileProxy.get_file_buffer(self.file_path)
            if self.file_buffer:
                # 2. Працюємо з Excel
                self.workbook = openpyxl.load_workbook(self.file_buffer, data_only=True)
                self.sheet = self.workbook[DESERTER_TAB_NAME]
                self._build_column_map()
                print(f'>> EXCEL LAST ROW::  {self._find_last_row()}')
        except Exception as e:
            # print(f"Помилка ініціалізації Excel: {e}")
            raise BaseException(f"⚠️ Помилка ініціалізації Excel: {e}")

    def save(self, fileProxy) -> None:
        if self.workbook is None:
            print("⚠️ Спроба зберегти порожній воркбук. Скасовано.")
            return
        try:
            output = io.BytesIO()
            self.workbook.save(output)
            size = output.tell()
            if size == 0:
                print("❌ Помилка: Openpyxl згенерував 0 байт даних!")
                return
            output.seek(0)
            with fileProxy as smb:
                smb.save_file_from_buffer(self.file_path, output)
            print(f"--- ✔️ EXCEL УСПІШНО ОНОВЛЕНО ({size} байт)")
        except Exception as e:
            print(f"❌ Критична помилка при збереженні: {e}")
        finally:
            output.close()
            if not self.batch_processing:
                self.workbook = None  # Очищуємо для наступних ітерацій

    def close(self):
        """Очищення ресурсів"""
        if self.workbook:
            self.workbook.close()
        self.workbook = None
        self.sheet = None