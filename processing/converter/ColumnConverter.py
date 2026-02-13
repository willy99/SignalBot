from dics.deserter_xls_dic import *
from openpyxl.styles import PatternFill
from processing.DocProcessor import DocProcessor
from processing.ExcelProcessor import ExcelProcessor
import traceback

class ColumnConverter:
    def __init__(self, excel_file_path):
        self.excelProcessor = ExcelProcessor(excel_file_path, batch_processing=True)
        self.docProcessor = DocProcessor(None, None, None)

    def convert(self):
        pass


    def _convert_subunit2(self):
        print("#convert")
        subunit_col = self.excelProcessor.column_map.get(COLUMN_SUBUNIT2.lower())
        bio_col = self.excelProcessor.column_map.get(COLUMN_BIO.lower())
        cond_col = self.excelProcessor.column_map.get(COLUMN_DESERT_CONDITIONS.lower())

        try:
            if not all([subunit_col, bio_col]):
                return None

            for row in range(2, self.excelProcessor.sheet.max_row + 1):
                s_bio = str(self.excelProcessor.sheet.cell(row=row, column=bio_col).value or "").strip()
                s_cond = str(self.excelProcessor.sheet.cell(row=row, column=cond_col).value or "").strip()
                cell = self.excelProcessor.sheet.cell(row=row, column=subunit_col)

                if (s_bio is None or s_bio.strip() == '') and (s_cond is None or s_cond == ''):
                    pale_red_fill = PatternFill(start_color='FFC7CE',
                                                end_color='FFC7CE',
                                                fill_type='solid')
                    cell.fill = pale_red_fill
                mil_subunit2 = self.docProcessor.extract_military_subunit(s_bio, None, PATTERN_SUBUNIT2_MAPPING)
                if mil_subunit2 is NA:
                    # fallback
                    mil_subunit2 = self.docProcessor.extract_military_subunit(s_cond, None, PATTERN_SUBUNIT2_MAPPING)
                cell.value = mil_subunit2

            self.excelProcessor.save(self.excelProcessor.fileProxy)
        except Exception as e:
            print(f"🔴 КРИТИЧНА ПОМИЛКА БАТЧУ: {e}")
            traceback.print_exc()
        finally:
            self.excelProcessor.close()
            print("🏁 >>> CONVERSION FINISHED")
