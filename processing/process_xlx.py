import openpyxl

def update_excel_status(file_path, search_text):
    # 1. Завантажуємо файл
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active  # Беремо перший лист

        found = False

        # 2. Проходимо по всіх ячейках для пошуку тексту
        # Ми перетворимо текст у строку і почистимо пробіли для надійності
        search_query = " ".join(str(search_text).lower().split())

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    # Очищаємо значення в ячейці від зайвих пробілів для порівняння
                    cell_value_clean = " ".join(str(cell.value).lower().split())

                    if search_query in cell_value_clean:
                        row_number = cell.row
                        # 3. Записуємо "+" у стовпець AQ (43-й за рахунком)
                        sheet.cell(row=row_number, column=43).value = "+"
                        print(f"Знайдено '{search_text}' у рядку {row_number}. Статус оновлено в AQ.")
                        found = True
                        break  # Виходимо з циклу по клітинках рядка
            if found:
                break  # Якщо потрібно оновити лише перше входження

        if not found:
            print(f"Текст '{search_text}' не знайдено.")
        else:
            # 4. Зберігаємо файл
            workbook.save(file_path)
            print("Файл успішно збережено.")

    except Exception as e:
        print(f"Виникла помилка: {e}")

